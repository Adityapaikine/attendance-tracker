import hashlib
import os
import tkinter as tk
from datetime import datetime, date as dt_date
from tkinter import ttk, messagebox, filedialog, simpledialog
from typing import Optional, List, Dict, Tuple
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from p2p_messenger import P2PMessenger
import re


# Professional Color Scheme
COLORS = {
    'primary': '#1e3a5f',      # Navy blue
    'secondary': '#4a90e2',    # Light blue
    'accent': '#2c5f8d',       # Medium blue
    'background': '#f5f7fa',   # Light gray-blue
    'surface': '#ffffff',      # White
    'text': '#2c3e50',         # Dark gray
    'text_light': '#7f8c8d',   # Light gray
    'success': '#27ae60',      # Green
    'warning': '#f39c12',      # Orange
    'error': '#e74c3c',         # Red
    'border': '#d5d8dc',        # Light border
}

THEMES = {
    "light": {
        "primary": "#1e3a5f",
        "secondary": "#4a90e2",
        "accent": "#2c5f8d",
        "background": "#f5f7fa",
        "surface": "#ffffff",
        "text": "#2c3e50",
        "text_light": "#7f8c8d",
        "success": "#27ae60",
        "warning": "#f39c12",
        "error": "#e74c3c",
        "border": "#d5d8dc",
    },

    "dark": {
        "primary": "#6ea8fe",
        "secondary": "#4d90fe",
        "accent": "#7aa2f7",
        "background": "#121212",
        "surface": "#1e1e1e",
        "text": "#f1f1f1",
        "text_light": "#a9a9a9",
        "success": "#4caf50",
        "warning": "#ffb74d",
        "error": "#ef5350",
        "border": "#3a3a3a",
    }
}

EXCEL_FILE_DEFAULT = "leaves_data_1.xlsx"
USERS_SHEET = "Users"
DEPARTMENTS_SHEET = "Departments"
LEAVES_SHEET = "Leaves"
HOLIDAYS_SHEET = "Holidays"
ATTENDANCE_SHEET = "Attendance"
ANNOUNCEMENTS_SHEET = "Announcements"


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def default_holidays_2026() -> List[List[str]]:
    # Matches the image you shared (same occasions/days, stored in ISO format).
    return [
        ["1", "2026-01-26", "Monday", "Republic Day"],
        ["2", "2026-03-03", "Tuesday", "Holi"],
        ["3", "2026-03-20", "Friday", "Ramzan"],
        ["4", "2026-04-03", "Friday", "Good Friday"],
        ["5", "2026-05-01", "Friday", "Indian Labour Day"],
        ["6", "2026-05-25", "Monday", "Memorial Day"],
        ["7", "2026-07-03", "Friday", "Independence Day (United States)"],
        ["8", "2026-09-07", "Monday", "US Labor Day"],
        ["9", "2026-09-25", "Friday", "Ganpati Visarjan"],
        ["10", "2026-10-02", "Friday", "Gandhi Jayanti"],
        ["11", "2026-11-09", "Monday", "Diwali"],
        ["12", "2026-11-26", "Thursday", "Thanksgiving Day"],
        ["13", "2026-11-27", "Friday", "After Thanksgiving day"],
    ]

def ensure_workbook(path: str) -> Tuple[
    Workbook,
    Worksheet,
    Worksheet,
    Worksheet,
    Worksheet,
    Worksheet,
    Worksheet,
    bool
]:
    """Ensure the Excel workbook exists and has the required sheets."""

    is_new = not os.path.exists(path)

    if is_new:

        wb = Workbook()

        # =====================================================
        # Users
        # =====================================================

        users_ws = wb.active
        users_ws.title = USERS_SHEET

        users_ws.append([
            "username",
            "full_name",
            "department",
            "designation",
            "dob",
            "phone_number",
            "email_id",
            "blood_group",
            "emergency_number",
            "password_hash",
            "role",
            "is_active"
        ])

        # =====================================================
        # Departments
        # =====================================================

        departments_ws = wb.create_sheet(
            DEPARTMENTS_SHEET
        )

        departments_ws.append([
            "department_id",
            "department_name",
            "is_active"
        ])

        # =====================================================
        # Leaves
        # =====================================================

        leaves_ws = wb.create_sheet(
            LEAVES_SHEET
        )

        leaves_ws.append([
            "leave_id",
            "username",
            "start_date",
            "end_date",
            "leave_type",
            "status",
            "reason",
            "created_at",
            "updated_at",
        ])

        # =====================================================
        # Holidays
        # =====================================================

        holidays_ws = wb.create_sheet(
            HOLIDAYS_SHEET
        )

        holidays_ws.append([
            "sr_no",
            "date",
            "day",
            "occasion"
        ])

        for holiday in default_holidays_2026():
            holidays_ws.append(holiday)

        # =====================================================
        # Attendance
        #
        # IMPORTANT:
        # New attendance format:
        #
        # Employee Name | User name | 01-Aug-26 | 02-Aug-26 ...
        #
        # =====================================================

        attendance_ws = wb.create_sheet(
            ATTENDANCE_SHEET
        )

        attendance_ws.append([
            "Employee Name",
            "User name"
        ])

        # =====================================================
        # Announcements
        # =====================================================

        announcements_ws = wb.create_sheet(
            ANNOUNCEMENTS_SHEET
        )

        announcements_ws.append([
            "announcement_id",
            "posted_by",
            "posted_at",
            "text",
            "photo_path",
        ])

        wb.save(path)

    else:

        wb = load_workbook(path)

        # =====================================================
        # Departments
        # =====================================================

        if DEPARTMENTS_SHEET not in wb.sheetnames:

            departments_ws = wb.create_sheet(
                DEPARTMENTS_SHEET
            )

            departments_ws.append([
                "department_id",
                "department_name",
                "is_active"
            ])

        else:

            departments_ws = wb[
                DEPARTMENTS_SHEET
            ]

        # =====================================================
        # Users
        # =====================================================

        if USERS_SHEET not in wb.sheetnames:

            users_ws = wb.create_sheet(
                USERS_SHEET
            )

            users_ws.append([
                "username",
                "full_name",
                "department",
                "password_hash",
                "role",
                "is_active"
            ])

        else:

            users_ws = wb[
                USERS_SHEET
            ]

            # Add department column if this is an older workbook
            headers = [
                cell.value
                for cell in users_ws[1]
            ]

            if "department" not in headers:

                users_ws.insert_cols(3)

                users_ws.cell(
                    1,
                    3
                ).value = "department"

                for row in range(
                    2,
                    users_ws.max_row + 1
                ):

                    users_ws.cell(
                        row,
                        3
                    ).value = ""

        # =====================================================
        # Leaves
        # =====================================================

        if LEAVES_SHEET not in wb.sheetnames:

            leaves_ws = wb.create_sheet(
                LEAVES_SHEET
            )

            leaves_ws.append([
                "leave_id",
                "username",
                "start_date",
                "end_date",
                "leave_type",
                "status",
                "reason",
                "created_at",
                "updated_at",
            ])

        else:

            leaves_ws = wb[
                LEAVES_SHEET
            ]

        # =====================================================
        # Holidays
        # =====================================================

        if HOLIDAYS_SHEET not in wb.sheetnames:

            holidays_ws = wb.create_sheet(
                HOLIDAYS_SHEET
            )

            holidays_ws.append([
                "sr_no",
                "date",
                "day",
                "occasion"
            ])

            for holiday in default_holidays_2026():
                holidays_ws.append(holiday)

        else:

            holidays_ws = wb[
                HOLIDAYS_SHEET
            ]

            if holidays_ws.max_row < 2:

                for holiday in default_holidays_2026():
                    holidays_ws.append(holiday)

        # =====================================================
        # Attendance
        # =====================================================

        if ATTENDANCE_SHEET not in wb.sheetnames:

            attendance_ws = wb.create_sheet(
                ATTENDANCE_SHEET
            )

            attendance_ws.append([
                "Employee Name",
                "User name"
            ])

        else:

            # IMPORTANT:
            # If Attendance already exists, DO NOT MODIFY IT.
            #
            # This preserves your existing attendance report:
            #
            # Employee Name | User name | 01-Aug-26 | ...
            #
            attendance_ws = wb[
                ATTENDANCE_SHEET
            ]

        # =====================================================
        # Announcements
        # =====================================================

        if ANNOUNCEMENTS_SHEET not in wb.sheetnames:

            announcements_ws = wb.create_sheet(
                ANNOUNCEMENTS_SHEET
            )

            announcements_ws.append([
                "announcement_id",
                "posted_by",
                "posted_at",
                "text",
                "photo_path",
            ])

        else:
            announcements_ws = wb[
                ANNOUNCEMENTS_SHEET
            ]

            # Repair an older/empty announcement sheet if necessary.
            headers = [
                str(cell.value).strip()
                if cell.value is not None else ""
                for cell in announcements_ws[1]
            ]

            required_headers = [
                "announcement_id",
                "posted_by",
                "posted_at",
                "text",
                "photo_path",
            ]

            if not headers or not all(
                header in headers
                for header in required_headers
            ):
                if announcements_ws.max_row <= 1:
                    for column, header in enumerate(
                        required_headers,
                        start=1
                    ):
                        announcements_ws.cell(
                            row=1,
                            column=column,
                            value=header
                        )

    return (
        wb,
        users_ws,
        departments_ws,
        leaves_ws,
        holidays_ws,
        attendance_ws,
        announcements_ws,
        is_new
    )


def load_users(users_ws: Worksheet) -> Dict[str, Dict]:
    users: Dict[str, Dict] = {}

    # Read the actual headers from Excel
    headers = []

    for cell in users_ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(
                str(cell.value).strip().lower()
            )

    # Map header name -> Excel column index
    header_index = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    def get_value(row, field, default=""):
        index = header_index.get(field)

        if index is None:
            return default

        if index >= len(row):
            return default

        value = row[index]

        if value is None:
            return default

        return value

    # Read every employee
    for row in users_ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row:
            continue

        username = get_value(
            row,
            "username"
        )

        if not username:
            continue

        username = str(
            username
        ).strip()

        users[username] = {

            "username": username,

            "full_name": str(
                get_value(
                    row,
                    "full_name"
                )
            ).strip(),

            "department": str(
                get_value(
                    row,
                    "department"
                )
            ).strip(),

            "password_hash": str(
                get_value(
                    row,
                    "password_hash"
                )
            ).strip(),

            "role": str(
                get_value(
                    row,
                    "role",
                    "user"
                )
            ).strip(),

            "is_active": str(
                get_value(
                    row,
                    "is_active",
                    True
                )
            ).strip().lower() != "false",

            # ---------------------------------------------
            # Profile information
            # ---------------------------------------------

            "designation": str(
                get_value(
                    row,
                    "designation"
                )
            ).strip(),

            "dob": str(
                get_value(
                    row,
                    "dob"
                )
            ).strip(),

            "phone_number": str(
                get_value(
                    row,
                    "phone_number"
                )
            ).strip(),

            "email_id": str(
                get_value(
                    row,
                    "email_id"
                )
            ).strip(),

            "blood_group": str(
                get_value(
                    row,
                    "blood_group"
                )
            ).strip(),

            "emergency_number": str(
                get_value(
                    row,
                    "emergency_number"
                )
            ).strip(),

            "photo_path": str(
                get_value(
                    row,
                    "photo_path"
                )
            ).strip(),

            "profile_completed": str(
                get_value(
                    row,
                    "profile_completed",
                    "FALSE"
                )
            ).strip().lower() == "true",
        }

    return users


def save_users(
    users_ws: Worksheet,
    users: Dict[str, Dict]
) -> None:

    required_headers = [
        "username",
        "full_name",
        "department",
        "password_hash",
        "role",
        "is_active",
        "designation",
        "dob",
        "phone_number",
        "email_id",
        "blood_group",
        "emergency_number",
        "photo_path",
        "profile_completed",
    ]

    # ---------------------------------------------------------
    # Make sure all required columns exist
    # ---------------------------------------------------------

    existing_headers = [
        str(cell.value).strip()
        if cell.value is not None
        else ""
        for cell in users_ws[1]
    ]

    for header in required_headers:

        if header not in existing_headers:

            users_ws.cell(
                row=1,
                column=users_ws.max_column + 1,
                value=header
            )

    # Re-read headers after adding missing columns
    headers = [
        str(cell.value).strip()
        if cell.value is not None
        else ""
        for cell in users_ws[1]
    ]

    header_index = {
        header: index + 1
        for index, header in enumerate(headers)
    }

    # ---------------------------------------------------------
    # Remove existing employee rows
    # ---------------------------------------------------------

    if users_ws.max_row > 1:

        users_ws.delete_rows(
            2,
            users_ws.max_row - 1
        )

    # ---------------------------------------------------------
    # Write users using header names
    # ---------------------------------------------------------

    for user in users.values():

        row_number = users_ws.max_row + 1

        values = {
            "username": user.get(
                "username",
                ""
            ),

            "full_name": user.get(
                "full_name",
                ""
            ),

            "department": user.get(
                "department",
                ""
            ),

            "password_hash": user.get(
                "password_hash",
                ""
            ),

            "role": user.get(
                "role",
                "user"
            ),

            "is_active": user.get(
                "is_active",
                True
            ),

            "designation": user.get(
                "designation",
                ""
            ),

            "dob": user.get(
                "dob",
                ""
            ),

            "phone_number": user.get(
                "phone_number",
                ""
            ),

            "email_id": user.get(
                "email_id",
                ""
            ),

            "blood_group": user.get(
                "blood_group",
                ""
            ),

            "emergency_number": user.get(
                "emergency_number",
                ""
            ),

            "photo_path": user.get(
                "photo_path",
                ""
            ),

            "profile_completed": user.get(
                "profile_completed",
                False
            ),

        }

        for field, value in values.items():

            column_number = header_index.get(field)

            if column_number is not None:

                users_ws.cell(
                    row=row_number,
                    column=column_number,
                    value=value
                )


def load_leaves(leaves_ws: Worksheet) -> List[Dict]:
    leaves: List[Dict] = []
    for row in leaves_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        (leave_id, username, start_date, end_date, leave_type, status, reason, created_at, updated_at) = row
        leaves.append({
            "leave_id": int(leave_id),
            "username": str(username),
            "start_date": str(start_date),
            "end_date": str(end_date),
            "leave_type": leave_type or "",
            "status": status or "PENDING",
            "reason": reason or "",
            "created_at": str(created_at) if created_at else "",
            "updated_at": str(updated_at) if updated_at else "",
        })
    return leaves


def save_leaves(leaves_ws: Worksheet, leaves: List[Dict]) -> None:
    leaves_ws.delete_rows(2, leaves_ws.max_row)
    for leave in leaves:
        leaves_ws.append([
            leave["leave_id"],
            leave["username"],
            leave["start_date"],
            leave["end_date"],
            leave["leave_type"],
            leave["status"],
            leave["reason"],
            leave["created_at"],
            leave["updated_at"],
        ])


def generate_next_leave_id(leaves: List[Dict]) -> int:
    if not leaves:
        return 1
    return max(leave["leave_id"] for leave in leaves) + 1


def load_holidays(holidays_ws: Worksheet) -> List[Dict]:
    holidays: List[Dict] = []
    for row in holidays_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sr_no, holiday_date, day, occasion = row
        normalized_date = ""
        if holiday_date:
            if isinstance(holiday_date, datetime):
                normalized_date = holiday_date.date().isoformat()
            elif isinstance(holiday_date, dt_date):
                normalized_date = holiday_date.isoformat()
            else:
                normalized_date = str(holiday_date)
        holidays.append({
            "sr_no": str(sr_no) if sr_no else "",
            "date": normalized_date,
            "day": str(day) if day else "",
            "occasion": str(occasion) if occasion else "",
        })
    return holidays


def save_holidays(holidays_ws: Worksheet, holidays: List[Dict]) -> None:
    holidays_ws.delete_rows(2, holidays_ws.max_row)
    for holiday in holidays:
        holidays_ws.append([
            holiday.get("sr_no", ""),
            holiday.get("date", ""),
            holiday.get("day", ""),
            holiday.get("occasion", ""),
        ])



def load_announcements(
    announcements_ws: Worksheet
) -> List[Dict]:
    """Load announcements and remove entries older than 7 days."""
    announcements: List[Dict] = []
    now = datetime.now()
    expired_ids = set()
    expired_photo_paths = []

    if announcements_ws.max_row < 2:
        return announcements

    headers = [
        str(cell.value).strip().lower()
        if cell.value is not None else ""
        for cell in announcements_ws[1]
    ]
    header_index = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    def get_value(row, field, default=""):
        index = header_index.get(field)
        if index is None or index >= len(row):
            return default
        return row[index] if row[index] is not None else default

    for row in announcements_ws.iter_rows(min_row=2, values_only=True):
        announcement_id = str(get_value(row, "announcement_id", "")).strip()
        if not announcement_id:
            continue

        posted_at_raw = get_value(row, "posted_at", "")
        posted_at = None

        if isinstance(posted_at_raw, datetime):
            posted_at = posted_at_raw
        elif posted_at_raw:
            try:
                posted_at = datetime.fromisoformat(
                    str(posted_at_raw).strip()
                )
            except ValueError:
                try:
                    posted_at = datetime.strptime(
                        str(posted_at_raw).strip(),
                        "%Y-%m-%d %H:%M:%S"
                    )
                except ValueError:
                    posted_at = None

        # If the timestamp is invalid, keep the announcement rather
        # than silently deleting it.
        if posted_at is not None:
            age_seconds = (now - posted_at).total_seconds()
            if age_seconds >= 7 * 24 * 60 * 60:
                expired_ids.add(announcement_id)
                photo_path = str(
                    get_value(row, "photo_path", "")
                ).strip()
                if photo_path:
                    expired_photo_paths.append(photo_path)
                continue

        announcements.append({
            "announcement_id": announcement_id,
            "posted_by": str(
                get_value(row, "posted_by", "")
            ).strip(),
            "posted_at": (
                posted_at.isoformat(sep=" ", timespec="seconds")
                if posted_at is not None
                else str(posted_at_raw)
            ),
            "text": str(
                get_value(row, "text", "")
            ).strip(),
            "photo_path": str(
                get_value(row, "photo_path", "")
            ).strip(),
        })

    # Physically remove expired rows from Excel.
    if expired_ids:
        save_announcements(
            announcements_ws,
            announcements
        )

        # Remove old announcement images if they are no longer referenced.
        referenced_photos = {
            str(a.get("photo_path", "")).strip()
            for a in announcements
            if a.get("photo_path")
        }

        for photo_path in expired_photo_paths:
            if (
                photo_path
                and photo_path not in referenced_photos
                and os.path.isfile(photo_path)
            ):
                try:
                    os.remove(photo_path)
                except OSError:
                    pass

    return announcements


def save_announcements(
    announcements_ws: Worksheet,
    announcements: List[Dict]
) -> None:
    """Save announcements to the Announcements worksheet."""
    if announcements_ws.max_row > 1:
        announcements_ws.delete_rows(
            2,
            announcements_ws.max_row - 1
        )

    for announcement in announcements:
        announcements_ws.append([
            announcement.get("announcement_id", ""),
            announcement.get("posted_by", ""),
            announcement.get("posted_at", ""),
            announcement.get("text", ""),
            announcement.get("photo_path", ""),
        ])


def generate_next_announcement_id(
    announcements: List[Dict]
) -> int:
    """Generate the next numeric announcement ID."""
    ids = []
    for announcement in announcements:
        try:
            ids.append(int(announcement.get("announcement_id", 0)))
        except (TypeError, ValueError):
            continue

    return max(ids, default=0) + 1


def load_attendance(
    att_ws: Worksheet,
    users: Optional[Dict[str, Dict]] = None
) -> List[Dict]:

    records: List[Dict] = []

    print("\n========== ATTENDANCE DEBUG ==========")
    print("Worksheet:", att_ws.title)
    print("Max rows:", att_ws.max_row)
    print("Max columns:", att_ws.max_column)

    print("Excel headers:")
    for cell in att_ws[1]:
        print(
            cell.column,
            repr(cell.value),
            type(cell.value)
        )

    print("First attendance row:")

    if att_ws.max_row >= 2:
        for cell in att_ws[2]:
            print(
                cell.column,
                repr(cell.value)
            )

    print("======================================\n")

    if att_ws.max_row < 2:
        return records

    headers = [
        cell.value
        for cell in att_ws[1]
    ]

    if len(headers) < 3:
        return records

    # ---------------------------------------------------------
    # YOUR FORMAT:
    #
    # Employee Name | User name | 01-Aug-26 | 02-Aug-26 | ...
    # ---------------------------------------------------------

    employee_header = str(
        headers[0] or ""
    ).strip().lower()

    username_header = str(
        headers[1] or ""
    ).strip().lower()

    if employee_header not in (
        "employee name",
        "employee",
        "name"
    ):
        return records

    if username_header not in (
        "user name",
        "username",
        "user"
    ):
        return records

    # ---------------------------------------------------------
    # Find all date columns
    # ---------------------------------------------------------

    date_columns = {}

    for col_index, header in enumerate(
        headers[2:],
        start=2
    ):

        if header is None:
            continue

        parsed_date = None

        if isinstance(header, datetime):
            parsed_date = header.date()

        elif isinstance(header, dt_date):
            parsed_date = header

        else:

            header_text = str(
                header
            ).strip()

            for fmt in (
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%d/%m/%y",
                "%Y-%m-%d",
                "%d-%B-%y",
                "%d-%B-%Y",
            ):

                try:

                    parsed_date = datetime.strptime(
                        header_text,
                        fmt
                    ).date()

                    break

                except ValueError:
                    continue

        if parsed_date:
            date_columns[col_index] = parsed_date

    # ---------------------------------------------------------
    # Read employee rows
    # ---------------------------------------------------------

    for row in att_ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row:
            continue

        employee_name = str(
            row[0] or ""
        ).strip()

        username = str(
            row[1] or ""
        ).strip()

        if not username:
            continue

        # -----------------------------------------------------
        # Read attendance for every date
        # -----------------------------------------------------

        for col_index, attendance_date in date_columns.items():

            if col_index >= len(row):
                continue

            raw_status = row[col_index]

            if raw_status is None:
                continue

            status_code = str(
                raw_status
            ).strip().upper()

            if not status_code:
                continue

            # -------------------------------------------------
            # Convert Excel codes
            # -------------------------------------------------

            status_map = {
                "P": "Present",
                "A": "Absent",
                "HD": "Half Day",
                "L": "Leave",
                "H": "Holiday",
                "WO": "Week Off",
                "WFH": "Work From Home",
                "PL": "Paid Leave",
                "ALWP": "Approved Leave Without Pay",
                "ULWP": "Unapproved Leave Without Pay",
                "-": "",
            }

            status = status_map.get(
                status_code,
                status_code
            )

            # Ignore "-"
            if not status:
                continue

            records.append({
                "date": attendance_date.isoformat(),
                "username": username,
                "employee_name": employee_name,
                "status": status,
                "remarks": "",
                "updated_at": "",
                "source_code": status_code,
            })

    return records

def save_attendance(
    att_ws: Worksheet,
    records: List[Dict]
) -> None:
    """
    Save attendance back to the Excel Attendance sheet.

    Excel format:

    Employee Name | User name | 01-Aug-26 | 02-Aug-26 | ...
    """

    # =========================================================
    # 1. Collect employee information
    # =========================================================

    employees = {}

    for record in records:

        username = str(
            record.get("username", "")
        ).strip()

        if not username:
            continue

        employee_name = str(
            record.get("employee_name", "")
        ).strip()

        # If employee name is missing, try Users data if available
        if not employee_name:
            employee_name = username

        employees[username] = employee_name

    # =========================================================
    # 2. Preserve existing employees from Excel
    # =========================================================

    existing_employees = {}

    for row in att_ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row:
            continue

        # Column A = Employee Name
        employee_name = str(
            row[0] or ""
        ).strip()

        # Column B = User name
        username = str(
            row[1] or ""
        ).strip() if len(row) > 1 else ""

        if username:
            existing_employees[username] = employee_name

    # Add existing employees that may not currently
    # have attendance records in self.attendance
    for username, employee_name in existing_employees.items():

        if username not in employees:
            employees[username] = employee_name

    # =========================================================
    # 3. Collect all attendance dates
    # =========================================================

    dates = set()

    # Dates from application records
    for record in records:

        date_value = record.get(
            "date",
            ""
        )

        if not date_value:
            continue

        try:

            parsed_date = datetime.strptime(
                str(date_value).strip(),
                "%Y-%m-%d"
            ).date()

            dates.add(parsed_date)

        except ValueError:
            continue

    # ---------------------------------------------------------
    # Also preserve dates already present in Excel
    # ---------------------------------------------------------

    for cell in att_ws[1][2:]:

        header = cell.value

        parsed_date = None

        if isinstance(header, datetime):

            parsed_date = header.date()

        elif isinstance(header, dt_date):

            parsed_date = header

        elif header:

            header_text = str(
                header
            ).strip()

            for fmt in (
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%d/%m/%y",
                "%Y-%m-%d",
                "%d-%B-%y",
                "%d-%B-%Y",
            ):

                try:

                    parsed_date = datetime.strptime(
                        header_text,
                        fmt
                    ).date()

                    break

                except ValueError:
                    continue

        if parsed_date:
            dates.add(parsed_date)

    # Sort dates chronologically
    dates = sorted(dates)

    # =========================================================
    # 4. Build attendance lookup
    # =========================================================
    #
    # (username, date) -> Excel status code
    #
    # =========================================================

    attendance_lookup = {}

    reverse_status_map = {
        "Present": "P",
        "Absent": "A",
        "Half Day": "HD",
        "Leave": "L",
        "Holiday": "H",
        "Week Off": "WO",
        "Work From Home": "WFH",

        # Preserve your attendance-system codes
        "PL": "PL",
        "ULWP": "ULWP",
        "ALWP": "ALWP",
    }

    for record in records:

        username = str(
            record.get("username", "")
        ).strip()

        date_value = str(
            record.get("date", "")
        ).strip()

        status = str(
            record.get("status", "")
        ).strip()

        source_code = str(
            record.get("source_code", "")
        ).strip().upper()

        if not username or not date_value:
            continue

        # -----------------------------------------------------
        # Prefer original Excel code when available
        # -----------------------------------------------------

        if source_code:
            excel_status = source_code

        else:
            excel_status = reverse_status_map.get(
                status,
                status
            )

        attendance_lookup[
            (username, date_value)
        ] = excel_status

    # =========================================================
    # 5. Clear existing attendance rows
    # =========================================================

    if att_ws.max_row > 1:

        att_ws.delete_rows(
            2,
            att_ws.max_row - 1
        )

    # =========================================================
    # 6. Write headers
    # =========================================================

    att_ws.cell(
        row=1,
        column=1,
        value="Employee Name"
    )

    att_ws.cell(
        row=1,
        column=2,
        value="User name"
    )

    # ---------------------------------------------------------
    # Write date headers
    # ---------------------------------------------------------

    for column_number, attendance_date in enumerate(
        dates,
        start=3
    ):

        cell = att_ws.cell(
            row=1,
            column=column_number,
            value=attendance_date
        )

        cell.number_format = "DD-MMM-YY"

    # =========================================================
    # 7. Write employee rows
    # =========================================================

    for row_number, username in enumerate(
        sorted(employees.keys()),
        start=2
    ):

        employee_name = employees.get(
            username,
            username
        )

        # -----------------------------------------------------
        # Column A = Employee Name
        # -----------------------------------------------------

        att_ws.cell(
            row=row_number,
            column=1,
            value=employee_name
        )

        # -----------------------------------------------------
        # Column B = User name
        # -----------------------------------------------------

        att_ws.cell(
            row=row_number,
            column=2,
            value=username
        )

        # -----------------------------------------------------
        # Attendance for each date
        # -----------------------------------------------------

        for column_number, attendance_date in enumerate(
            dates,
            start=3
        ):

            date_iso = attendance_date.isoformat()

            status = attendance_lookup.get(
                (
                    username,
                    date_iso
                ),
                ""
            )

            att_ws.cell(
                row=row_number,
                column=column_number,
                value=status
            )

def load_attendance(
    att_ws: Worksheet,
    users: Optional[Dict[str, Dict]] = None
) -> List[Dict]:
    """
    Load attendance from Excel.

    Excel format:

    Employee Name | User name | 01-Aug-26 | 02-Aug-26 | ...
    """

    records: List[Dict] = []

    if att_ws.max_row < 2:
        return records

    headers = [
        cell.value
        for cell in att_ws[1]
    ]

    if len(headers) < 3:
        return records

    # ---------------------------------------------------------
    # Validate headers
    # ---------------------------------------------------------

    employee_header = str(
        headers[0] or ""
    ).strip().lower()

    username_header = str(
        headers[1] or ""
    ).strip().lower()

    if employee_header not in (
        "employee name",
        "employee",
        "name"
    ):
        return records

    if username_header not in (
        "user name",
        "username",
        "user"
    ):
        return records

    # ---------------------------------------------------------
    # Find date columns
    # ---------------------------------------------------------

    date_columns = {}

    for col_index, header in enumerate(
        headers[2:],
        start=2
    ):

        if header is None:
            continue

        parsed_date = None

        if isinstance(header, datetime):
            parsed_date = header.date()

        elif isinstance(header, dt_date):
            parsed_date = header

        else:

            header_text = str(
                header
            ).strip()

            for fmt in (
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%d/%m/%y",
                "%Y-%m-%d",
                "%d-%B-%y",
                "%d-%B-%Y",
            ):
                try:
                    parsed_date = datetime.strptime(
                        header_text,
                        fmt
                    ).date()
                    break
                except ValueError:
                    continue

        if parsed_date:
            date_columns[col_index] = parsed_date

    # ---------------------------------------------------------
    # Read employee rows
    # ---------------------------------------------------------

    for row in att_ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row:
            continue

        employee_name = str(
            row[0] or ""
        ).strip()

        username = str(
            row[1] or ""
        ).strip()

        if not username:
            continue

        # -----------------------------------------------------
        # Read every attendance date
        # -----------------------------------------------------

        for col_index, attendance_date in date_columns.items():

            if col_index >= len(row):
                continue

            raw_status = row[col_index]

            if raw_status is None:
                continue

            status_code = str(
                raw_status
            ).strip().upper()

            if not status_code or status_code == "-":
                continue

            # -------------------------------------------------
            # Excel code -> application status
            # -------------------------------------------------

            status_map = {
                "P": "Present",
                "A": "Absent",
                "HD": "Half Day",
                "L": "Leave",
                "H": "Holiday",
                "WO": "Week Off",
                "WFH": "Work From Home",
                "PL": "Paid Leave",
                "ALWP": "Approved Leave Without Pay",
                "ULWP": "Unapproved Leave Without Pay",
            }

            status = status_map.get(
                status_code,
                status_code
            )

            records.append({
                "date": attendance_date.isoformat(),
                "username": username,
                "employee_name": employee_name,
                "status": status,
                "remarks": "",
                "updated_at": "",
                "source_code": status_code,
            })

    return records








def load_departments(departments_ws: Worksheet) -> List[Dict]:
    departments = []

    for row in departments_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        department_id, department_name, is_active = row

        departments.append({
            "department_id": int(department_id),
            "department_name": str(department_name),
            "is_active": str(is_active).lower() != "false",
        })

    return departments


def save_departments(
    departments_ws: Worksheet,
    departments: List[Dict]
) -> None:

    departments_ws.delete_rows(2, departments_ws.max_row)

    for department in departments:
        departments_ws.append([
            department["department_id"],
            department["department_name"],
            department.get("is_active", True),
        ])


def generate_next_department_id(departments: List[Dict]) -> int:
    if not departments:
        return 1

    return max(
        d["department_id"]
        for d in departments
    ) + 1

class LeaveTrackerApp:

    def configure_treeview_theme(self):
        """Configure Treeview colors for the current application theme."""

        style = ttk.Style(self.root)

        try:
            style.theme_use("clam")
        except Exception:
            pass

        if self.theme == "dark":

            style.configure(
                "Treeview",
                background="#1e1e1e",
                foreground="#f1f1f1",
                fieldbackground="#1e1e1e",
                bordercolor="#3a3a3a",
                borderwidth=1,
                rowheight=28
            )

            style.configure(
                "Treeview.Heading",
                background="#252525",
                foreground="#ffffff",
                relief="flat",
                font=("Segoe UI", 10, "bold")
            )

            style.map(
                "Treeview",
                background=[
                    ("selected", "#4d90fe")
                ],
                foreground=[
                    ("selected", "#ffffff")
                ]
            )

            style.map(
                "Treeview.Heading",
                background=[
                    ("active", "#333333")
                ],
                foreground=[
                    ("active", "#ffffff")
                ]
            )

        else:

            style.configure(
                "Treeview",
                background="#ffffff",
                foreground="#222222",
                fieldbackground="#ffffff",
                bordercolor="#d5d8dc",
                borderwidth=1,
                rowheight=28
            )

            style.configure(
                "Treeview.Heading",
                background="#f0f0f0",
                foreground="#222222",
                relief="flat",
                font=("Segoe UI", 10, "bold")
            )

            style.map(
                "Treeview",
                background=[
                    ("selected", "#4a90e2")
                ],
                foreground=[
                    ("selected", "#ffffff")
                ]
            )

    def toggle_theme(self):
        self.theme = (
            "dark"
            if self.theme == "light"
            else "light"
        )

        COLORS.clear()
        COLORS.update(
            THEMES[self.theme]
        )

        self.refresh_theme()

    def refresh_theme(self):
        """Refresh the entire application for the selected theme."""

        # -------------------------------------------------
        # Root window
        # -------------------------------------------------
        self.root.configure(
            bg=COLORS["background"]
        )
        
        # -------------------------------------------------
        # Main application frame
        # -------------------------------------------------
        if hasattr(self, "main_frame"):
            self.main_frame.configure(
                bg=COLORS["background"]
            )
        self.configure_treeview_theme()

        # -------------------------------------------------
        # Rebuild the current page
        # -------------------------------------------------
        if self.current_user:

            role = self.current_user.get(
                "role",
                ""
            ).lower()

            if role == "master":
                self.show_master_dashboard()
            else:
                self.show_user_dashboard()

        else:
            self.show_login_screen()

        # -------------------------------------------------
        # Update theme button text
        # -------------------------------------------------
        if hasattr(self, "theme_button"):

            if self.theme == "dark":

                self.theme_button.configure(
                    text="☀ Light Mode",
                    bg=COLORS["surface"],
                    fg=COLORS["text"],
                    activebackground=COLORS["background"],
                    activeforeground=COLORS["text"]
                )

            else:

                self.theme_button.configure(
                    text="🌙 Dark Mode",
                    bg=COLORS["surface"],
                    fg=COLORS["text"],
                    activebackground=COLORS["background"],
                    activeforeground=COLORS["text"]
                )
    

    def handle_new_message_notification(
        self,
        event=None
    ):
        """
        Display a notification when a new message arrives.
        """

        try:

            if not hasattr(
                self,
                "messenger"
            ):

                return

            unread = (
                self.messenger.get_unread_count()
            )

            if unread <= 0:
                return

            # -----------------------------------------------------
            # Update application title
            # -----------------------------------------------------

            if unread == 1:

                title = (
                    "Leave Management System "
                    "— 💬 1 New Message"
                )

            else:

                title = (
                    "Leave Management System "
                    f"— 💬 {unread} New Messages"
                )

            self.root.title(title)

            # -----------------------------------------------------
            # Show popup
            # -----------------------------------------------------

            messagebox.showinfo(
                "💬 New Message",
                (
                    f"You have {unread} unread "
                    "message"
                    + (
                        ""
                        if unread == 1
                        else "s"
                    )
                    + ".\n\n"
                    "Open Messages to read them."
                )
            )

        except Exception as e:

            print(
                "[DASHBOARD MESSAGE NOTIFICATION ERROR]",
                e
            )
    

    def on_closing(self):
        """Cleanly close the application."""

        try:
            if hasattr(self, "messenger") and self.messenger:
                self.messenger.stop_network()

        except Exception as e:
            print(
                "[MESSENGER SHUTDOWN ERROR]",
                e
            )

        self.root.destroy()
    
    def open_messenger(self):
        """Open the integrated P2P LAN messenger."""

        if not hasattr(
            self,
            "messenger"
        ) or self.messenger is None:

            messagebox.showerror(
                "Messages",
                "Messaging service is not running."
            )

            return

        self.messenger.open_window()

    def show_department_employees_box(self, parent):
        """
        Display employees and their details based on the logged-in user.

        MASTER:
            Can see all departments and choose a department.

        NORMAL USER:
            Can see only employees belonging to the same department
            as the logged-in user.
        """

        # =========================================================
        # SECURITY
        # =========================================================

        if not self.current_user:
            return None

        current_role = str(
            self.current_user.get("role", "")
        ).strip().lower()

        current_department = str(
            self.current_user.get("department", "")
        ).strip()

        is_master = (
            current_role == "master"
        )

        # =========================================================
        # CARD
        # =========================================================

        card = tk.Frame(
            parent,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        card.pack(
            fill=tk.BOTH,
            expand=True,
            padx=25,
            pady=(10, 15)
        )

        # =========================================================
        # HEADER
        # =========================================================

        header = tk.Frame(
            card,
            bg=COLORS["surface"]
        )

        header.pack(
            fill=tk.X,
            padx=18,
            pady=(15, 8)
        )

        tk.Label(
            header,
            text="Department Employees",
            font=("Segoe UI", 15, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["primary"]
        ).pack(
            side=tk.LEFT
        )

        # =========================================================
        # DEPARTMENT INFORMATION / FILTER
        # =========================================================

        filter_frame = tk.Frame(
            card,
            bg=COLORS["surface"]
        )

        filter_frame.pack(
            fill=tk.X,
            padx=18,
            pady=(0, 10)
        )

        tk.Label(
            filter_frame,
            text="Department:",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["text"]
        ).pack(
            side=tk.LEFT,
            padx=(0, 8)
        )

        department_var = tk.StringVar()

        # ---------------------------------------------------------
        # MASTER
        # ---------------------------------------------------------

        if is_master:

            department_names = [
                str(
                    d.get(
                        "department_name",
                        ""
                    )
                ).strip()
                for d in self.departments
                if d.get("is_active", True)
                and d.get("department_name")
            ]

            department_names = sorted(
                set(department_names),
                key=lambda x: x.lower()
            )

            department_values = [
                "All Departments"
            ] + department_names

            department_combo = ttk.Combobox(
                filter_frame,
                textvariable=department_var,
                values=department_values,
                state="readonly",
                width=30
            )

            department_combo.pack(
                side=tk.LEFT
            )

            department_var.set(
                "All Departments"
            )

        # ---------------------------------------------------------
        # NORMAL USER
        # ---------------------------------------------------------

        else:

            # The user cannot change this department.
            department_var.set(
                current_department
            )

            tk.Label(
                filter_frame,
                text=current_department or "No Department",
                font=("Segoe UI", 10, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["secondary"]
            ).pack(
                side=tk.LEFT
            )

        # =========================================================
        # TABLE
        # =========================================================

        table_frame = tk.Frame(
            card,
            bg=COLORS["surface"]
        )

        table_frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=18,
            pady=(0, 12)
        )

        columns = (
            "Name",
            "Department",
            "Designation",
            "Phone",
            "Email",
            "Blood Group"
        )

        tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=8
        )

        for column in columns:

            tree.heading(
                column,
                text=column
            )

        tree.column(
            "Name",
            width=180,
            anchor=tk.W
        )

        tree.column(
            "Department",
            width=140,
            anchor=tk.W
        )

        tree.column(
            "Designation",
            width=160,
            anchor=tk.W
        )

        tree.column(
            "Phone",
            width=130,
            anchor=tk.CENTER
        )

        tree.column(
            "Email",
            width=230,
            anchor=tk.W
        )

        tree.column(
            "Blood Group",
            width=100,
            anchor=tk.CENTER
        )

        scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.VERTICAL,
            command=tree.yview
        )

        tree.configure(
            yscrollcommand=scrollbar.set
        )

        tree.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True
        )

        scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y
        )

        # =========================================================
        # GET VISIBLE EMPLOYEES
        # =========================================================

        def get_visible_employees():

            employees = [
                user
                for user in self.users.values()
                if (
                    user.get("role") == "user"
                    and user.get("is_active", True)
                )
            ]

            # -----------------------------------------------------
            # MASTER
            # -----------------------------------------------------

            if is_master:

                selected_department = (
                    department_var.get()
                    .strip()
                )

                if selected_department != "All Departments":

                    employees = [
                        user
                        for user in employees
                        if str(
                            user.get(
                                "department",
                                ""
                            )
                        ).strip().lower()
                        ==
                        selected_department.lower()
                    ]

            # -----------------------------------------------------
            # NORMAL USER
            # -----------------------------------------------------

            else:

                # IMPORTANT:
                # Do not trust a UI selection for normal users.
                # Always use the department stored against the
                # logged-in account.
                employees = [
                    user
                    for user in employees
                    if str(
                        user.get(
                            "department",
                            ""
                        )
                    ).strip().lower()
                    ==
                    current_department.lower()
                ]

            return sorted(
                employees,
                key=lambda user: (
                    user.get("full_name")
                    or user.get("username")
                    or ""
                ).lower()
            )

        # =========================================================
        # REFRESH
        # =========================================================

        def refresh_employees(*_):

            tree.delete(
                *tree.get_children()
            )

            employees = (
                get_visible_employees()
            )

            for user in employees:

                username = str(
                    user.get(
                        "username",
                        ""
                    )
                ).strip()

                tree.insert(
                    "",
                    tk.END,
                    iid=username,
                    values=(
                        user.get(
                            "full_name",
                            ""
                        ),

                        user.get(
                            "department",
                            ""
                        ),

                        user.get(
                            "designation",
                            ""
                        ),

                        user.get(
                            "phone_number",
                            ""
                        ),

                        user.get(
                            "email_id",
                            ""
                        ),

                        user.get(
                            "blood_group",
                            ""
                        )
                    )
                )

            if not employees:

                tree.insert(
                    "",
                    tk.END,
                    values=(
                        "No employees found",
                        "",
                        "",
                        "",
                        "",
                        ""
                    )
                )

        # =========================================================
        # MASTER DEPARTMENT CHANGE
        # =========================================================

        if is_master:

            department_combo.bind(
                "<<ComboboxSelected>>",
                refresh_employees
            )

        # =========================================================
        # VIEW EMPLOYEE DETAILS
        # =========================================================

        def view_selected_employee():

            selected = tree.selection()

            if not selected:

                messagebox.showwarning(
                    "Select Employee",
                    "Please select an employee first."
                )

                return

            username = selected[0]

            user = self.users.get(
                username
            )

            if not user:

                messagebox.showerror(
                    "Error",
                    "Employee information could not be found."
                )

                return

            # -----------------------------------------------------
            # SECURITY CHECK
            # -----------------------------------------------------

            if not is_master:

                employee_department = str(
                    user.get(
                        "department",
                        ""
                    )
                ).strip()

                if (
                    employee_department.lower()
                    != current_department.lower()
                ):

                    messagebox.showerror(
                        "Access Denied",
                        "You can only view employees "
                        "from your own department."
                    )

                    return

            # =====================================================
            # DETAILS WINDOW
            # =====================================================

            dialog = tk.Toplevel(
                self.root
            )

            dialog.title(
                "Employee Details"
            )

            dialog.geometry(
                "650x650"
            )

            dialog.configure(
                bg=COLORS["background"]
            )

            dialog.transient(
                self.root
            )

            tk.Label(
                dialog,
                text="Employee Details",
                font=("Segoe UI", 20, "bold"),
                bg=COLORS["background"],
                fg=COLORS["primary"]
            ).pack(
                pady=(20, 15)
            )

            details_frame = tk.Frame(
                dialog,
                bg=COLORS["surface"],
                highlightbackground=COLORS["border"],
                highlightthickness=1
            )

            details_frame.pack(
                fill=tk.BOTH,
                expand=True,
                padx=30,
                pady=10
            )

            # =====================================================
            # PHOTO
            # =====================================================

            photo_label = tk.Label(
                details_frame,
                text="No Photo",
                font=("Segoe UI", 11, "bold"),
                width=16,
                height=8,
                bg="#eeeeee",
                fg=COLORS["text_light"],
                relief=tk.SOLID,
                bd=1
            )

            photo_label.pack(
                pady=(20, 15)
            )

            photo_path = str(
                user.get(
                    "photo_path",
                    ""
                )
            ).strip()

            if photo_path and os.path.isfile(
                photo_path
            ):

                try:

                    image = Image.open(
                        photo_path
                    )

                    image.thumbnail(
                        (150, 150)
                    )

                    photo = ImageTk.PhotoImage(
                        image
                    )

                    photo_label.config(
                        image=photo,
                        text=""
                    )

                    photo_label.image = photo

                except Exception:
                    pass

            # =====================================================
            # DETAILS
            # =====================================================

            fields = [
                ("Name", "full_name"),
                ("Department", "department"),
                ("Designation", "designation"),
                ("DOB", "dob"),
                ("Phone Number", "phone_number"),
                ("Email ID", "email_id"),
                ("Blood Group", "blood_group"),
                ("Emergency Number", "emergency_number"),
            ]

            details = tk.Frame(
                details_frame,
                bg=COLORS["surface"]
            )

            details.pack(
                fill=tk.X,
                padx=30
            )

            for row_index, (
                label_text,
                field_name
            ) in enumerate(fields):

                tk.Label(
                    details,
                    text=label_text + ":",
                    font=("Segoe UI", 10, "bold"),
                    bg=COLORS["surface"],
                    fg=COLORS["text"],
                    anchor=tk.W
                ).grid(
                    row=row_index,
                    column=0,
                    sticky="w",
                    padx=5,
                    pady=6
                )

                tk.Label(
                    details,
                    text=str(
                        user.get(
                            field_name,
                            ""
                        )
                    ),
                    font=("Segoe UI", 10),
                    bg=COLORS["surface"],
                    fg=COLORS["text"],
                    anchor=tk.W
                ).grid(
                    row=row_index,
                    column=1,
                    sticky="w",
                    padx=15,
                    pady=6
                )

            details.columnconfigure(
                1,
                weight=1
            )

            tk.Button(
                dialog,
                text="Close",
                font=("Segoe UI", 11),
                bg=COLORS["text_light"],
                fg=COLORS["surface"],
                relief=tk.FLAT,
                padx=25,
                pady=8,
                command=dialog.destroy
            ).pack(
                pady=15
            )

        # =========================================================
        # BUTTON
        # =========================================================

        #tk.Button(
            card,
            text="View Selected Employee Details",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["secondary"],
            fg=COLORS["surface"],
            activebackground=COLORS["accent"],
            activeforeground=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=20,
            pady=7,
            command=view_selected_employee
        #).pack(
            pady=(0, 15)
        #)

        # =========================================================
        # INITIAL LOAD
        # =========================================================

        refresh_employees()

        return card

    def show_all_employee_profiles(self):
        
        #"""Master-only view of all employee profile details."""

        # ---------------------------------------------------------
        # Security check
        # ---------------------------------------------------------

        if not self.current_user:
            messagebox.showerror(
                "Access Denied",
                "You must be logged in."
            )
            return

        if self.current_user.get("role") != "master":
            messagebox.showerror(
                "Access Denied",
                "Only the Super User can view all employee profiles."
            )
            return

        # ---------------------------------------------------------
        # Create dialog
        # ---------------------------------------------------------

        dialog = tk.Toplevel(self.root)
        dialog.title("All Employee Profiles")
        dialog.geometry("1250x600")
        dialog.configure(
            bg=COLORS["background"]
        )

        dialog.transient(self.root)

        # ---------------------------------------------------------
        # Header
        # ---------------------------------------------------------

        tk.Label(
            dialog,
            text="All Employee Profiles",
            font=("Segoe UI", 18, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(
            pady=(15, 5)
        )

        tk.Label(
            dialog,
            text="Master User View",
            font=("Segoe UI", 10),
            bg=COLORS["background"],
            fg=COLORS["text_light"]
        ).pack(
            pady=(0, 10)
        )

        # ---------------------------------------------------------
        # Table container
        # ---------------------------------------------------------

        table_frame = tk.Frame(
            dialog,
            bg=COLORS["surface"]
        )

        table_frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=10
        )

        columns = (
            "Name",
            "Designation",
            "DOB",
            "Phone Number",
            "Email ID",
            "Blood Group",
            "Emergency Number"
        )

        tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings"
        )

        # ---------------------------------------------------------
        # Column headings
        # ---------------------------------------------------------

        for column in columns:

            tree.heading(
                column,
                text=column
            )

        # ---------------------------------------------------------
        # Column widths
        # ---------------------------------------------------------

        tree.column(
            "Name",
            width=180,
            anchor=tk.W
        )

        tree.column(
            "Designation",
            width=170,
            anchor=tk.W
        )

        tree.column(
            "DOB",
            width=110,
            anchor=tk.CENTER
        )

        tree.column(
            "Phone Number",
            width=140,
            anchor=tk.CENTER
        )

        tree.column(
            "Email ID",
            width=230,
            anchor=tk.W
        )

        tree.column(
            "Blood Group",
            width=110,
            anchor=tk.CENTER
        )

        tree.column(
            "Emergency Number",
            width=160,
            anchor=tk.CENTER
        )

        # ---------------------------------------------------------
        # Scrollbars
        # ---------------------------------------------------------

        vertical_scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.VERTICAL,
            command=tree.yview
        )

        horizontal_scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.HORIZONTAL,
            command=tree.xview
        )

        tree.configure(
            yscrollcommand=vertical_scrollbar.set,
            xscrollcommand=horizontal_scrollbar.set
        )

        tree.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True
        )

        vertical_scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y
        )

        horizontal_scrollbar.pack(
            side=tk.BOTTOM,
            fill=tk.X
        )

        # ---------------------------------------------------------
        # Load all employees
        # ---------------------------------------------------------

        employee_users = [
            user
            for user in self.users.values()
            if user.get("role") == "user"
            and user.get("is_active", True)
        ]

        employee_users.sort(
            key=lambda user: (
                user.get("full_name")
                or user.get("username")
                or ""
            ).lower()
        )

        for user in employee_users:

            tree.insert(
                "",
                tk.END,
                values=(
                    user.get("full_name", ""),
                    user.get("designation", ""),
                    user.get("dob", ""),
                    user.get("phone_number", ""),
                    user.get("email_id", ""),
                    user.get("blood_group", ""),
                    user.get("emergency_number", "")
                )
            )

        # ---------------------------------------------------------
        # No employees
        # ---------------------------------------------------------

        if not employee_users:

            tk.Label(
                dialog,
                text="No active employees found.",
                font=("Segoe UI", 11),
                bg=COLORS["background"],
                fg=COLORS["text_light"]
            ).pack(
                pady=10
            )

        # ---------------------------------------------------------
        # Close button
        # ---------------------------------------------------------

        tk.Button(
            dialog,
            text="Close",
            font=("Segoe UI", 11),
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            command=dialog.destroy,
            padx=25,
            pady=8
        ).pack(
            pady=15
        )

    def show_my_profile(self):
        """Display and manage the logged-in employee profile."""

        if not self.current_user:
            messagebox.showerror(
                "Error",
                "No user is currently logged in."
            )
            return

        username = self.current_user["username"]

        user = self.users.get(
            username,
            self.current_user
        )

        is_master = (
            user.get("role") == "master"
        )

        profile_completed = user.get(
            "profile_completed",
            False
        )

        # ---------------------------------------------------------
        # Window
        # ---------------------------------------------------------

        dialog = tk.Toplevel(self.root)

        dialog.title("My Profile")
        dialog.geometry("750x760")
        dialog.configure(
            bg=COLORS["background"]
        )

        dialog.transient(self.root)

        # ---------------------------------------------------------
        # Header
        # ---------------------------------------------------------

        header = tk.Frame(
            dialog,
            bg=COLORS["primary"],
            height=80
        )

        header.pack(
            fill=tk.X
        )

        header.pack_propagate(False)

        tk.Label(
            header,
            text="My Profile",
            font=("Segoe UI", 22, "bold"),
            bg=COLORS["primary"],
            fg=COLORS["surface"]
        ).pack(
            pady=22
        )

        # ---------------------------------------------------------
        # Main container
        # ---------------------------------------------------------

        main = tk.Frame(
            dialog,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        main.pack(
            fill=tk.BOTH,
            expand=True,
            padx=30,
            pady=20
        )

        # =========================================================
        # PHOTO
        # =========================================================

        photo_frame = tk.Frame(
            main,
            bg=COLORS["surface"]
        )

        photo_frame.pack(
            pady=15
        )

        photo_label = tk.Label(
            photo_frame,
            text="No Photo",
            font=("Segoe UI", 12, "bold"),
            width=16,
            height=8,
            bg="#eeeeee",
            fg=COLORS["text_light"],
            relief=tk.SOLID,
            bd=1
        )

        photo_label.pack()

        current_photo = {
            "path": user.get(
                "photo_path",
                ""
            )
        }

        # ---------------------------------------------------------
        # Display photo
        # ---------------------------------------------------------

        def display_photo(path):

            if not path:
                photo_label.config(
                    image="",
                    text="No Photo"
                )
                photo_label.image = None
                return

            if not os.path.exists(path):
                photo_label.config(
                    image="",
                    text="Photo\nnot found"
                )
                photo_label.image = None
                return

            try:
                # Tkinter PhotoImage supports PNG/GIF.
                # PIL is recommended for JPG/JPEG.
                from PIL import Image, ImageTk

                image = Image.open(path)

                image.thumbnail(
                    (150, 150)
                )

                photo = ImageTk.PhotoImage(
                    image
                )

                photo_label.config(
                    image=photo,
                    text=""
                )

                photo_label.image = photo

            except Exception:
                photo_label.config(
                    image="",
                    text="Unable to\nload photo"
                )

                photo_label.image = None

        display_photo(
            current_photo["path"]
        )

        # =========================================================
        # PHOTO BUTTON
        # =========================================================

        def choose_photo():

            path = filedialog.askopenfilename(
                title="Select Profile Photo",
                filetypes=[
                    (
                        "Image Files",
                        "*.jpg *.jpeg *.png *.gif"
                    ),
                    (
                        "JPEG",
                        "*.jpg *.jpeg"
                    ),
                    (
                        "PNG",
                        "*.png"
                    ),
                    (
                        "GIF",
                        "*.gif"
                    ),
                ]
            )

            if not path:
                return

            current_photo["path"] = path

            display_photo(
                path
            )

        tk.Button(
            main,
            text="Change / Add Photo",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["secondary"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=20,
            pady=7,
            command=choose_photo
        ).pack(
            pady=(0, 15)
        )

        # =========================================================
        # PROFILE STATUS
        # =========================================================

        if is_master:

            status_text = (
                "Master User: Profile details can be edited anytime."
            )

        elif profile_completed:

            status_text = (
                "Profile completed. Personal details are locked. "
                "You can still change your photo."
            )

        else:

            status_text = (
                "Please complete your profile. "
                "After saving, your personal details will be locked."
            )

        tk.Label(
            main,
            text=status_text,
            font=("Segoe UI", 10),
            bg=COLORS["surface"],
            fg=COLORS["text_light"],
            wraplength=650
        ).pack(
            pady=(0, 15)
        )

        # =========================================================
        # PROFILE FIELDS
        # =========================================================

        profile_frame = tk.Frame(
            main,
            bg=COLORS["surface"]
        )

        profile_frame.pack(
            fill=tk.X,
            padx=25
        )

        fields = [
            ("Name", "full_name"),
            ("Designation", "designation"),
            ("DOB", "dob"),
            ("Phone Number", "phone_number"),
            ("Email ID", "email_id"),
            ("Blood Group", "blood_group"),
            ("Emergency Number", "emergency_number"),
        ]

        entries = {}

        # ---------------------------------------------------------
        # Determine whether fields are editable
        # ---------------------------------------------------------

        can_edit_details = (
            is_master
            or not profile_completed
        )

        for row_index, (
            label_text,
            field_name
        ) in enumerate(fields):

            tk.Label(
                profile_frame,
                text=label_text + ":",
                font=("Segoe UI", 10, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["text"],
                anchor=tk.W
            ).grid(
                row=row_index,
                column=0,
                sticky="w",
                padx=10,
                pady=7
            )

            entry = tk.Entry(
                profile_frame,
                font=("Segoe UI", 10),
                width=42,
                relief=tk.SOLID,
                bd=1
            )

            entry.insert(
                0,
                str(
                    user.get(
                        field_name,
                        ""
                    )
                )
            )

            entry.grid(
                row=row_index,
                column=1,
                sticky="ew",
                padx=10,
                pady=7
            )

            if not can_edit_details:

                entry.config(
                    state="readonly",
                    readonlybackground="#eeeeee"
                )

            entries[field_name] = entry

        profile_frame.columnconfigure(
            1,
            weight=1
        )

        # =========================================================
        # SAVE
        # =========================================================

        def save_profile():

            # -----------------------------------------------------
            # Always allow photo update
            # -----------------------------------------------------

            user["photo_path"] = (
                current_photo["path"]
            )

            # -----------------------------------------------------
            # Master can always edit details
            # -----------------------------------------------------

            if is_master:

                for field_name in entries:

                    user[field_name] = (
                        entries[field_name]
                        .get()
                        .strip()
                    )

            # -----------------------------------------------------
            # Employee can edit only if profile isn't completed
            # -----------------------------------------------------

            elif not profile_completed:

                for field_name in entries:

                    user[field_name] = (
                        entries[field_name]
                        .get()
                        .strip()
                    )

                # Lock profile permanently for employee
                user["profile_completed"] = True

            # -----------------------------------------------------
            # Employee with completed profile
            # -----------------------------------------------------

            else:

                # Do NOT modify any personal details.
                # Only photo is saved.

                pass

            # -----------------------------------------------------
            # Synchronize current user
            # -----------------------------------------------------

            self.current_user = user

            # -----------------------------------------------------
            # Save
            # -----------------------------------------------------

            if self.save_data():

                if is_master:

                    message = (
                        "Profile updated successfully."
                    )

                elif profile_completed:

                    message = (
                        "Profile photo updated successfully."
                    )

                else:

                    message = (
                        "Profile completed successfully.\n\n"
                        "Your personal details are now locked. "
                        "You can still change your photo."
                    )

                messagebox.showinfo(
                    "Profile",
                    message
                )

                dialog.destroy()

        # =========================================================
        # BUTTONS
        # =========================================================

        button_frame = tk.Frame(
            dialog,
            bg=COLORS["background"]
        )

        button_frame.pack(
            pady=(0, 20)
        )

        tk.Button(
            button_frame,
            text=(
                "Save Profile"
                if can_edit_details
                else "Save Photo"
            ),
            font=("Segoe UI", 11, "bold"),
            bg=COLORS["success"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=25,
            pady=9,
            command=save_profile
        ).pack(
            side=tk.LEFT,
            padx=8
        )

        tk.Button(
            button_frame,
            text="Close",
            font=("Segoe UI", 11),
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=25,
            pady=9,
            command=dialog.destroy
        ).pack(
            side=tk.LEFT,
            padx=8
        )

        


    def __init__(self, root):
        self.root = root
        self.theme = "light"
        self.root.bind(
            "<<MessengerNewMessage>>",
            self.handle_new_message_notification
        )
        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.on_closing
        )
        self.root.title("Leave Management System")
        self.root.geometry("1000x700")
        self.root.configure(bg=COLORS['background'])

        # Data storage
        self.excel_path = r"C:\Users\Aditya\Desktop\Code\leaves_data_1.xlsx"
        self.wb = None
        self.users_ws = None
        self.leaves_ws = None
        self.holidays_ws = None
        self.attendance_ws = None
        self.announcements_ws = None
        self.users = {}
        self.departments = []
        self.leaves = []
        self.holidays = []
        self.attendance: List[Dict] = []
        self.announcements: List[Dict] = []
        self.current_user = None
        
        self.setup_ui()
        self.root.bind(
            "<<MessengerNewMessage>>",
            self.handle_new_message_notification
        )
        self.load_data()



    def show_manage_departments(self):
    #"""Master-only department management."""

        if self.current_user.get("role") != "master":
            messagebox.showerror(
                "Access Denied",
                "Only the Super User can manage departments."
            )
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Manage Departments")
        dialog.geometry("650x450")
        dialog.configure(bg=COLORS["background"])

        tk.Label(
            dialog,
            text="Manage Departments",
            font=("Segoe UI", 16, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(pady=15)

        tree = ttk.Treeview(
            dialog,
            columns=("ID", "Department", "Status"),
            show="headings",
            height=14
        )

        tree.heading("ID", text="ID")
        tree.heading("Department", text="Department")
        tree.heading("Status", text="Status")

        tree.column("ID", width=70, anchor=tk.CENTER)
        tree.column("Department", width=300)
        tree.column("Status", width=120, anchor=tk.CENTER)

        tree.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=10
        )

        def refresh():
            tree.delete(*tree.get_children())

            for d in self.departments:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        d["department_id"],
                        d["department_name"],
                        "Active" if d.get("is_active", True)
                        else "Inactive"
                    )
                )

        def add_department():
            name = tk.simpledialog.askstring(
                "Add Department",
                "Department name:"
            )

            if not name:
                return

            name = name.strip()

            if any(
                d["department_name"].lower() == name.lower()
                for d in self.departments
            ):
                messagebox.showerror(
                    "Error",
                    "Department already exists."
                )
                return

            self.departments.append({
                "department_id":
                    generate_next_department_id(self.departments),
                "department_name": name,
                "is_active": True,
            })

            if self.save_data():
                refresh()

        def edit_department():
            selected = tree.selection()

            if not selected:
                messagebox.showwarning(
                    "Select",
                    "Please select a department."
                )
                return

            values = tree.item(selected[0])["values"]
            department_id = int(values[0])

            department = next(
                (
                    d for d in self.departments
                    if d["department_id"] == department_id
                ),
                None
            )

            if not department:
                return

            name = tk.simpledialog.askstring(
                "Edit Department",
                "Department name:",
                initialvalue=department["department_name"]
            )

            if not name:
                return

            name = name.strip()

            department["department_name"] = name

            if self.save_data():
                refresh()

        def deactivate_department():
            selected = tree.selection()

            if not selected:
                messagebox.showwarning(
                    "Select",
                    "Please select a department."
                )
                return

            values = tree.item(selected[0])["values"]
            department_id = int(values[0])

            department = next(
                (
                    d for d in self.departments
                    if d["department_id"] == department_id
                ),
                None
            )

            if not department:
                return

            department["is_active"] = False

            if self.save_data():
                refresh()

        buttons = tk.Frame(
            dialog,
            bg=COLORS["background"]
        )

        buttons.pack(pady=10)

        tk.Button(
            buttons,
            text="Add",
            bg=COLORS["success"],
            fg=COLORS["surface"],
            command=add_department,
            padx=20,
            pady=7
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            buttons,
            text="Edit",
            bg=COLORS["secondary"],
            fg=COLORS["surface"],
            command=edit_department,
            padx=20,
            pady=7
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            buttons,
            text="Deactivate",
            bg=COLORS["error"],
            fg=COLORS["surface"],
            command=deactivate_department,
            padx=20,
            pady=7
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            buttons,
            text="Close",
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            command=dialog.destroy,
            padx=20,
            pady=7
        ).pack(side=tk.LEFT, padx=5)

        refresh()

    def toggle_theme(self):
        self.theme = (
            "dark"
            if self.theme == "light"
            else "light"
        )

        COLORS.clear()
        COLORS.update(
            THEMES[self.theme]
        )

        self.refresh_theme()
        
    def setup_ui(self):
        """Setup the main UI components."""
        
        # Header Frame
        header_frame = tk.Frame(self.root, bg=COLORS['primary'], height=80)
        header_frame.pack(fill=tk.X, padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="Leave Management System",
            font=('Segoe UI', 24, 'bold'),
            bg=COLORS['primary'],
            fg=COLORS['surface'],
            pady=20
        )
        title_label.pack()
        
        # Main container
        self.main_frame = tk.Frame(self.root, bg=COLORS['background'])
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Show login screen initially
        self.show_login_screen()

        
        
    def load_data(self):
        """Load data from Excel file."""

        try:

            import os

            print("========================================")
            print("ACTUAL EXCEL PATH")
            print("========================================")
            print("self.excel_path =", repr(self.excel_path))
            print("absolute path    =", os.path.abspath(self.excel_path))
            print("exists           =", os.path.exists(self.excel_path))

            if os.path.exists(self.excel_path):
                print(
                    "file size        =",
                    os.path.getsize(self.excel_path),
                    "bytes"
                )

            print("========================================")
            (
                self.wb,
                self.users_ws,
                self.departments_ws,
                self.leaves_ws,
                self.holidays_ws,
                self.attendance_ws,
                self.announcements_ws,
                is_new,
            ) = ensure_workbook(self.excel_path)

            # =====================================================
            # DEBUG - Excel file and Attendance sheet
            # =====================================================

            print("\n========================================")
            print("ATTENDANCE DEBUG")
            print("========================================")

            print("Excel file:")
            print(self.excel_path)

            print("\nWorkbook sheets:")
            print(self.wb.sheetnames)

            print("\nAttendance worksheet:")
            print(self.attendance_ws)

            if self.attendance_ws is not None:

                print(
                    "Attendance sheet name:",
                    self.attendance_ws.title
                )

                print(
                    "Rows:",
                    self.attendance_ws.max_row
                )

                print(
                    "Columns:",
                    self.attendance_ws.max_column
                )

                print("\nAttendance headers:")

                for cell in self.attendance_ws[1]:
                    print(
                        cell.column,
                        repr(cell.value),
                        type(cell.value)
                    )

                if self.attendance_ws.max_row >= 2:

                    print("\nFirst employee row:")

                    for cell in self.attendance_ws[2]:
                        print(
                            cell.column,
                            repr(cell.value)
                        )

            print("========================================\n")

            # =====================================================
            # Load normal data
            # =====================================================

            self.users = load_users(
                self.users_ws
            )

            self.departments = load_departments(
                self.departments_ws
            )

            self.leaves = load_leaves(
                self.leaves_ws
            )

            self.holidays = load_holidays(
                self.holidays_ws
            )

            # =====================================================
            # Load attendance
            # =====================================================

            self.attendance = load_attendance(
                self.attendance_ws,
                self.users
            )

            # Load announcements and automatically purge anything
            # that has reached 7 days from posting.
            self.announcements = load_announcements(
                self.announcements_ws
            )

            # Persist the automatic 7-day cleanup immediately.
            if self.announcements_ws is not None:
                save_announcements(
                    self.announcements_ws,
                    self.announcements
                )
                self.wb.save(self.excel_path)

            print(
                "ATTENDANCE RECORDS LOADED:",
                len(self.attendance)
            )

            print(
                "ANNOUNCEMENTS LOADED:",
                len(self.announcements)
            )

            if self.attendance:

                print(
                    "FIRST ATTENDANCE RECORD:"
                )

                print(
                    self.attendance[0]
                )

                print(
                    "LAST ATTENDANCE RECORD:"
                )

                print(
                    self.attendance[-1]
                )

            else:

                print(
                    "WARNING: NO ATTENDANCE RECORDS LOADED!"
                )

            # =====================================================
            # Holidays
            # =====================================================

            if (
                self.holidays_ws is not None
                and not self.holidays
            ):

                for holiday in default_holidays_2026():
                    self.holidays_ws.append(holiday)

                self.wb.save(
                    self.excel_path
                )

                self.holidays = load_holidays(
                    self.holidays_ws
                )

            # =====================================================
            # Check master user
            # =====================================================

            has_master = any(
                u["role"] == "master"
                and u["is_active"]
                for u in self.users.values()
            )

            if not has_master and not is_new:
                self.show_master_creation_dialog()

        except Exception as e:

            import traceback

            print("\n========== LOAD ERROR ==========")

            traceback.print_exc()

            print("================================\n")

            messagebox.showerror(
                "Error",
                f"Failed to load data:\n\n{str(e)}"
            )
            
    def save_data(self):
        """Save data to Excel file."""
        try:
            save_users(self.users_ws, self.users)
            save_departments(self.departments_ws, self.departments)
            save_leaves(self.leaves_ws, self.leaves)
            if self.holidays_ws is not None:
                save_holidays(self.holidays_ws, self.holidays)
            if self.announcements_ws is not None:
                save_announcements(
                    self.announcements_ws,
                    self.announcements
                )
            self.wb.save(self.excel_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")
            return False
            
    def show_login_screen(self):
        """Display login screen."""
        self.clear_main_frame()
        
        login_frame = tk.Frame(self.main_frame, bg=COLORS['surface'], relief=tk.RAISED, bd=2)
        login_frame.pack(expand=True, fill=tk.BOTH, padx=100, pady=50)
        
        # Title
        title = tk.Label(
            login_frame,
            text="Login",
            font=('Segoe UI', 20, 'bold'),
            bg=COLORS['surface'],
            fg=COLORS['primary'],
            pady=30
        )
        title.pack()
        
        # Username
        tk.Label(login_frame, text="User ID:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=10)
        self.username_entry = tk.Entry(login_frame, font=('Segoe UI', 11), width=30, relief=tk.SOLID, bd=1)
        self.username_entry.pack(pady=5)
        
        # Password
        tk.Label(login_frame, text="Password:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=10)
        self.password_entry = tk.Entry(login_frame, font=('Segoe UI', 11), width=30, show="*", relief=tk.SOLID, bd=1)
        self.password_entry.pack(pady=5)
        
        # Login button
        login_btn = tk.Button(
            login_frame,
            text="Login",
            font=('Segoe UI', 12, 'bold'),
            bg=COLORS['secondary'],
            fg=COLORS['surface'],
            activebackground=COLORS['accent'],
            activeforeground=COLORS['surface'],
            relief=tk.FLAT,
            cursor='hand2',
            padx=30,
            pady=10,
            command=self.handle_login
        )
        login_btn.pack(pady=30)
        
        self.username_entry.focus()
        self.root.bind('<Return>', lambda e: self.handle_login())
        
        
    def handle_login(self):
        """Handle login attempt."""
        username = self.username_entry.get().strip()
        password = self.password_entry.get()

        user = self.users.get(username)

        if not user or not user.get("is_active", True):
            messagebox.showerror(
                "Login Failed",
                "Invalid credentials or inactive user."
            )
            return

        if user.get("password_hash") != hash_password(password):
            messagebox.showerror(
                "Login Failed",
                "Invalid credentials."
            )
            return

        self.current_user = user

        # =====================================================
        # Start P2P Messenger
        # =====================================================

        try:

            if hasattr(
                self,
                "messenger"
            ):

                self.messenger.stop_network()

            self.messenger = P2PMessenger(
                root=self.root,
                username=self.current_user["username"],
                users_provider=lambda: self.users,
                colors=COLORS
            )

        except Exception as e:

            print(
                "[MESSENGER START ERROR]",
                e
            )

        messagebox.showinfo(
            "Success",
            f"Welcome, {user.get('full_name') or user['username']}!"
        )

        if user["role"] == "master":

            self.show_master_dashboard()

        else:

            self.show_user_dashboard()
            
    def _purge_expired_announcements(self, save=True):
        """Remove announcements that are 7 days old or older."""
        now = datetime.now()
        active = []
        expired_photo_paths = []

        for announcement in self.announcements:
            posted_at_raw = str(
                announcement.get("posted_at", "")
            ).strip()

            posted_at = None
            try:
                posted_at = datetime.fromisoformat(posted_at_raw)
            except ValueError:
                try:
                    posted_at = datetime.strptime(
                        posted_at_raw,
                        "%Y-%m-%d %H:%M:%S"
                    )
                except ValueError:
                    pass

            if posted_at is not None:
                if (now - posted_at).total_seconds() >= (
                    7 * 24 * 60 * 60
                ):
                    photo_path = str(
                        announcement.get("photo_path", "")
                    ).strip()

                    if photo_path:
                        expired_photo_paths.append(
                            photo_path
                        )

                    continue

            active.append(announcement)

        if len(active) != len(self.announcements):
            self.announcements = active

            if save and self.announcements_ws is not None:
                try:
                    save_announcements(
                        self.announcements_ws,
                        self.announcements
                    )
                    self.wb.save(self.excel_path)
                except Exception:
                    # Do not make the dashboard unusable because
                    # an automatic cleanup save failed.
                    pass

        referenced_photos = {
            str(a.get("photo_path", "")).strip()
            for a in self.announcements
            if a.get("photo_path")
        }

        for photo_path in expired_photo_paths:
            if (
                photo_path
                and photo_path not in referenced_photos
                and os.path.isfile(photo_path)
            ):
                try:
                    os.remove(photo_path)
                except OSError:
                    pass

    def show_announcements_section(
        self,
        parent,
        master_controls=True,
        grid_position=None
    ):
        """Display the shared announcement section on a dashboard."""
        self._purge_expired_announcements()

        card = tk.Frame(
            parent,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        if grid_position is not None:
            card.grid(
                row=grid_position[0],
                column=grid_position[1],
                sticky="nsew",
                padx=(12, 0),
                pady=0
            )
        else:
            card.pack(
                fill=tk.BOTH,
                expand=True,
                padx=25,
                pady=(10, 15)
            )

        header = tk.Frame(
            card,
            bg=COLORS["surface"]
        )
        header.pack(
            fill=tk.X,
            padx=18,
            pady=(15, 10)
        )

        tk.Label(
            header,
            text="Announcements",
            font=("Segoe UI", 15, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["primary"]
        ).pack(side=tk.LEFT)

        tk.Label(
            header,
            text="Posts are automatically removed after 7 days",
            font=("Segoe UI", 9),
            bg=COLORS["surface"],
            fg=COLORS["text_light"]
        ).pack(side=tk.LEFT, padx=12)

        if (
            master_controls
            and self.current_user
            and self.current_user.get("role") == "master"
        ):
            tk.Button(
                header,
                text="Post Announcement",
                font=("Segoe UI", 10, "bold"),
                bg=COLORS["secondary"],
                fg=COLORS["surface"],
                activebackground=COLORS["accent"],
                activeforeground=COLORS["surface"],
                relief=tk.FLAT,
                cursor="hand2",
                padx=15,
                pady=6,
                command=self.show_post_announcement_dialog
            ).pack(side=tk.RIGHT)

        # Scrollable announcement area.
        canvas = tk.Canvas(
            card,
            bg=COLORS["background"],
            highlightthickness=0
        )

        scrollbar = ttk.Scrollbar(
            card,
            orient=tk.VERTICAL,
            command=canvas.yview
        )

        list_frame = tk.Frame(
            canvas,
            bg=COLORS["background"]
        )

        canvas_window = canvas.create_window(
            (0, 0),
            window=list_frame,
            anchor="nw"
        )

        canvas.configure(
            yscrollcommand=scrollbar.set
        )

        self.theme_button = tk.Button(
        header,
        text="🌙 Dark Mode",
        command=self.toggle_theme,
        bg=COLORS["surface"],
        fg=COLORS["text"],
        activebackground=COLORS["background"],
        activeforeground=COLORS["text"],
        relief="flat",
        cursor="hand2",
        font=("Segoe UI", 10, "bold")
        )

        self.theme_button.pack(
            side="right",
            padx=10,
            pady=8
        )


        def on_frame_configure(_event=None):
            canvas.configure(
                scrollregion=canvas.bbox("all")
            )

        def on_canvas_configure(event):
            canvas.itemconfigure(
                canvas_window,
                width=event.width
            )

        list_frame.bind(
            "<Configure>",
            on_frame_configure
        )

        canvas.bind(
            "<Configure>",
            on_canvas_configure
        )

        canvas.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True,
            padx=(10, 0),
            pady=(0, 10)
        )

        scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y,
            pady=(0, 10),
            padx=(0, 10)
        )

        if not self.announcements:
            tk.Label(
                list_frame,
                text="No active announcements.",
                font=("Segoe UI", 11),
                bg=COLORS["background"],
                fg=COLORS["text_light"]
            ).pack(
                pady=35
            )
            return card

        sorted_announcements = sorted(
            self.announcements,
            key=lambda a: a.get("posted_at", ""),
            reverse=True
        )

        for announcement in sorted_announcements:
            announcement_card = tk.Frame(
                list_frame,
                bg=COLORS["surface"],
                highlightbackground=COLORS["border"],
                highlightthickness=1
            )
            announcement_card.pack(
                fill=tk.X,
                padx=10,
                pady=7
            )

            top = tk.Frame(
                announcement_card,
                bg=COLORS["surface"]
            )
            top.pack(
                fill=tk.X,
                padx=12,
                pady=(10, 4)
            )

            posted_at = str(
                announcement.get("posted_at", "")
            ).strip()

            try:
                posted_display = datetime.fromisoformat(
                    posted_at
                ).strftime("%d-%b-%Y %I:%M %p")
            except ValueError:
                posted_display = posted_at

            tk.Label(
                top,
                text=(
                    f"Posted by: "
                    f"{announcement.get('posted_by', '')}   "
                    f"|   {posted_display}"
                ),
                font=("Segoe UI", 9, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["text_light"]
            ).pack(side=tk.LEFT)

            content = tk.Frame(
                announcement_card,
                bg=COLORS["surface"]
            )
            content.pack(
                fill=tk.X,
                padx=12,
                pady=(0, 12)
            )

            text_value = str(
                announcement.get("text", "")
            ).strip()

            if text_value:
                tk.Label(
                    content,
                    text=text_value,
                    font=("Segoe UI", 11),
                    bg=COLORS["surface"],
                    fg=COLORS["text"],
                    justify=tk.LEFT,
                    anchor=tk.W,
                    wraplength=430
                ).pack(
                    side=tk.LEFT,
                    fill=tk.X,
                    expand=True,
                    padx=(0, 12)
                )

            photo_path = str(
                announcement.get("photo_path", "")
            ).strip()

            if photo_path and os.path.isfile(photo_path):
                try:
                    image = Image.open(photo_path)
                    image.thumbnail((220, 150))
                    photo = ImageTk.PhotoImage(image)

                    image_label = tk.Label(
                        content,
                        image=photo,
                        bg=COLORS["surface"]
                    )
                    image_label.image = photo
                    image_label.pack(
                        side=tk.RIGHT,
                        padx=(5, 0)
                    )
                except Exception:
                    pass

        return card

    def show_post_announcement_dialog(self):
        """Master-only dialog to create a new announcement."""
        if not self.current_user:
            messagebox.showerror(
                "Access Denied",
                "You must be logged in."
            )
            return

        if self.current_user.get("role") != "master":
            messagebox.showerror(
                "Access Denied",
                "Only the Super User can post announcements."
            )
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Post Announcement")
        dialog.geometry("700x620")
        dialog.configure(bg=COLORS["background"])
        dialog.transient(self.root)

        tk.Label(
            dialog,
            text="Post Announcement",
            font=("Segoe UI", 18, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(pady=(18, 4))

        tk.Label(
            dialog,
            text="The announcement will automatically expire 7 days after posting.",
            font=("Segoe UI", 10),
            bg=COLORS["background"],
            fg=COLORS["text_light"]
        ).pack(pady=(0, 12))

        form = tk.Frame(
            dialog,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )
        form.pack(
            fill=tk.BOTH,
            expand=True,
            padx=25,
            pady=10
        )

        tk.Label(
            form,
            text="Announcement Text:",
            font=("Segoe UI", 11, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["text"]
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(18, 6)
        )

        text_box = tk.Text(
            form,
            height=9,
            wrap=tk.WORD,
            font=("Segoe UI", 11),
            relief=tk.SOLID,
            bd=1
        )
        text_box.pack(
            fill=tk.BOTH,
            expand=True,
            padx=18,
            pady=(0, 15)
        )

        selected_photo = {"path": ""}

        photo_info = tk.Label(
            form,
            text="No photo selected.",
            font=("Segoe UI", 9),
            bg=COLORS["surface"],
            fg=COLORS["text_light"],
            anchor=tk.W
        )
        photo_info.pack(
            fill=tk.X,
            padx=18,
            pady=(0, 8)
        )

        def choose_photo():
            path = filedialog.askopenfilename(
                title="Select Announcement Photo",
                filetypes=[
                    (
                        "Image Files",
                        "*.jpg *.jpeg *.png *.gif *.bmp *.webp"
                    ),
                    (
                        "JPEG",
                        "*.jpg *.jpeg"
                    ),
                    (
                        "PNG",
                        "*.png"
                    ),
                    (
                        "All Files",
                        "*.*"
                    ),
                ]
            )

            if path:
                selected_photo["path"] = path
                photo_info.config(
                    text=f"Selected: {path}"
                )

        tk.Button(
            form,
            text="Choose Photo (Optional)",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["secondary"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=15,
            pady=7,
            command=choose_photo
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(0, 18)
        )

        def post():
            announcement_text = text_box.get(
                "1.0",
                tk.END
            ).strip()

            source_photo = selected_photo["path"]

            if not announcement_text and not source_photo:
                messagebox.showwarning(
                    "Announcement",
                    "Please enter announcement text or select a photo."
                )
                return

            photo_path = ""

            # Copy uploaded image to a folder next to the Excel file
            # so the announcement keeps working after the source file
            # is moved or the application is restarted.
            if source_photo:
                try:
                    from shutil import copy2

                    image_folder = os.path.join(
                        os.path.dirname(
                            os.path.abspath(
                                self.excel_path
                            )
                        ),
                        "announcement_images"
                    )

                    os.makedirs(
                        image_folder,
                        exist_ok=True
                    )

                    extension = os.path.splitext(
                        source_photo
                    )[1].lower()

                    safe_extension = (
                        extension
                        if extension in (
                            ".jpg",
                            ".jpeg",
                            ".png",
                            ".gif",
                            ".bmp",
                            ".webp",
                        )
                        else ".jpg"
                    )

                    file_name = (
                        f"announcement_"
                        f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
                        f"{safe_extension}"
                    )

                    photo_path = os.path.join(
                        image_folder,
                        file_name
                    )

                    copy2(
                        source_photo,
                        photo_path
                    )

                except Exception as e:
                    messagebox.showerror(
                        "Photo Error",
                        f"Could not save the announcement photo:\n\n{e}"
                    )
                    return

            announcement = {
                "announcement_id": str(
                    generate_next_announcement_id(
                        self.announcements
                    )
                ),
                "posted_by": (
                    self.current_user.get("full_name")
                    or self.current_user.get("username", "")
                ),
                "posted_at": datetime.now().isoformat(
                    sep=" ",
                    timespec="seconds"
                ),
                "text": announcement_text,
                "photo_path": photo_path,
            }

            self.announcements.insert(
                0,
                announcement
            )

            if self.save_data():
                messagebox.showinfo(
                    "Announcement",
                    "Announcement posted successfully."
                )
                dialog.destroy()

                # Refresh the dashboard immediately.
                if self.current_user.get("role") == "master":
                    self.show_master_dashboard()
                else:
                    self.show_user_dashboard()

        buttons = tk.Frame(
            dialog,
            bg=COLORS["background"]
        )
        buttons.pack(
            pady=(0, 18)
        )

        tk.Button(
            buttons,
            text="Post",
            font=("Segoe UI", 11, "bold"),
            bg=COLORS["success"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=25,
            pady=8,
            command=post
        ).pack(
            side=tk.LEFT,
            padx=6
        )

        tk.Button(
            buttons,
            text="Cancel",
            font=("Segoe UI", 11),
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=25,
            pady=8,
            command=dialog.destroy
        ).pack(
            side=tk.LEFT,
            padx=6
        )

    def get_message_unread_count(self):

        try:

            if hasattr(
                self,
                "messenger"
            ) and self.messenger:

                return self.messenger.get_unread_count()

        except Exception:
            pass

        return 0

    def show_master_dashboard(self):
        
        """Display a professional Master User dashboard."""
        self.clear_main_frame()

        # =====================================================
        # Dashboard header
        # =====================================================

        header = tk.Frame(
            self.main_frame,
            bg=COLORS["background"]
        )

        header.pack(
            fill=tk.X,
            padx=25,
            pady=(5, 15)
        )

        # =====================================================
        # Department Employees
        # =====================================================

        self.show_department_employees_box(
            self.main_frame
        )

        # -----------------------------------------------------
        # Dashboard title
        # -----------------------------------------------------

        tk.Label(
            header,
            text="Master Dashboard",
            font=("Segoe UI", 22, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(
            side=tk.LEFT
        )

        # -----------------------------------------------------
        # Logged-in user information
        # -----------------------------------------------------

        user_info_frame = tk.Frame(
            header,
            bg=COLORS["background"]
        )

        user_info_frame.pack(
            side=tk.RIGHT,
            padx=10
        )

        # Name
        tk.Label(
            user_info_frame,
            text=(
                f"Welcome, "
                f"{self.current_user.get('full_name') or self.current_user['username']}"
            ),
            font=("Segoe UI", 11, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"],
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )

        # Designation
        tk.Label(
            user_info_frame,
            text=(
                f"Designation: "
                f"{self.current_user.get('designation') or 'N/A'}"
            ),
            font=("Segoe UI", 9),
            bg=COLORS["background"],
            fg=COLORS["text_light"],
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )

        # Status
        employee_status = (
            "Active"
            if self.current_user.get("is_active", True)
            else "Inactive"
        )

        tk.Label(
            user_info_frame,
            text=f"Status: {employee_status}",
            font=("Segoe UI", 9, "bold"),
            bg=COLORS["background"],
            fg=(
                COLORS["success"]
                if employee_status == "Active"
                else COLORS["error"]
            ),
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )














        
        # =====================================================
        # Two-column dashboard
        # =====================================================
        content = tk.Frame(
            self.main_frame,
            bg=COLORS["background"]
        )
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 10))

        content.columnconfigure(0, weight=0, minsize=315)
        content.columnconfigure(1, weight=1)
        content.rowconfigure(0, weight=1)

        # =====================================================
        # Left: Administration menu
        # =====================================================

        menu_card = tk.Frame(
            content,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        menu_card.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 12)
        )

        # =====================================================
        # Administration Header
        # =====================================================

        tk.Label(
            menu_card,
            text="Administration",
            font=("Segoe UI", 14, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["primary"]
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(18, 3)
        )

        tk.Label(
            menu_card,
            text="Manage employees, attendance and leave",
            font=("Segoe UI", 9),
            bg=COLORS["surface"],
            fg=COLORS["text_light"]
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(0, 10)
        )

        # =====================================================
        # Scrollable Administration Area
        # =====================================================

        admin_container = tk.Frame(
            menu_card,
            bg=COLORS["surface"]
        )

        admin_container.pack(
            fill=tk.BOTH,
            expand=True,
            padx=8,
            pady=(0, 8)
        )

        # =====================================================
        # Canvas
        # =====================================================

        admin_canvas = tk.Canvas(
            admin_container,
            bg=COLORS["surface"],
            highlightthickness=0,
            bd=0
        )

        admin_canvas.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True
        )

        # =====================================================
        # Scrollbar
        # =====================================================

        admin_scrollbar = ttk.Scrollbar(
            admin_container,
            orient=tk.VERTICAL,
            command=admin_canvas.yview
        )

        admin_scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y
        )

        admin_canvas.configure(
            yscrollcommand=admin_scrollbar.set
        )

        # =====================================================
        # Inner scrolling frame
        # =====================================================

        admin_buttons_frame = tk.Frame(
            admin_canvas,
            bg=COLORS["surface"]
        )

        admin_window = admin_canvas.create_window(
            (0, 0),
            window=admin_buttons_frame,
            anchor="nw"
        )

        # =====================================================
        # Update scroll region
        # =====================================================

        def update_admin_scroll_region(event=None):

            admin_canvas.configure(
                scrollregion=admin_canvas.bbox("all")
            )

        admin_buttons_frame.bind(
            "<Configure>",
            update_admin_scroll_region
        )

        # =====================================================
        # Keep inner frame width equal to canvas width
        # =====================================================

        def update_admin_width(event):

            admin_canvas.itemconfigure(
                admin_window,
                width=event.width
            )

        admin_canvas.bind(
            "<Configure>",
            update_admin_width
        )

        # =====================================================
        # Mouse Wheel
        # =====================================================

        def admin_mousewheel(event):

            admin_canvas.yview_scroll(
                int(-1 * (event.delta / 120)),
                "units"
            )

        admin_canvas.bind(
            "<Enter>",
            lambda event: admin_canvas.bind_all(
                "<MouseWheel>",
                admin_mousewheel
            )
        )

        admin_canvas.bind(
            "<Leave>",
            lambda event: admin_canvas.unbind_all(
                "<MouseWheel"
            )
        )

        # =====================================================
        # Administration Buttons
        # =====================================================

        buttons = [
            ("Add User", self.show_add_user_dialog),
            ("Manage Users", self.show_manage_users),
            ("Employee Profiles", self.show_all_employee_profiles),
            ("Manage Departments", self.show_manage_departments),
            ("Approve Leaves", self.show_approve_leaves),
            ("Manage Holidays", self.show_manage_holidays),
            ("Daily Attendance", self.show_daily_attendance),
            ("Attendance Summary", self.show_attendance_summary),
            ("View All Leaves", self.show_all_leaves),
            ("💬 Messages", self.open_messenger),
        ]

        for text, command in buttons:

            btn = tk.Button(
                admin_buttons_frame,
                text=text,
                font=("Segoe UI", 10, "bold"),
                bg=COLORS["secondary"],
                fg=COLORS["surface"],
                activebackground=COLORS["accent"],
                activeforeground=COLORS["surface"],
                relief=tk.FLAT,
                cursor="hand2",
                anchor=tk.W,
                padx=16,
                pady=8,
                command=command
            )

            btn.pack(
                fill=tk.X,
                padx=8,
                pady=4
            )

        # =====================================================
        # Separator
        # =====================================================

        tk.Frame(
            admin_buttons_frame,
            bg=COLORS["border"],
            height=1
        ).pack(
            fill=tk.X,
            padx=8,
            pady=(12, 10)
        )

        # =====================================================
        # Logout
        # =====================================================

        tk.Button(
            admin_buttons_frame,
            text="Logout",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["error"],
            fg=COLORS["surface"],
            activebackground="#c0392b",
            activeforeground=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            padx=16,
            pady=9,
            command=self.logout
        ).pack(
            fill=tk.X,
            padx=8,
            pady=(0, 12)
        )
                

        # =====================================================
        # Right: Announcements
        # =====================================================
        self.show_announcements_section(
            content,
            master_controls=True,
            grid_position=(0, 1)
        )

    def _parse_date_input(self, text: str) -> Optional[dt_date]:
        s = (text or "").strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            return None

    def _reindex_holidays(self):
        for idx, h in enumerate(self.holidays, start=1):
            h["sr_no"] = str(idx)

    def _get_attendance_record(self, date_iso: str, username: str) -> Optional[Dict]:
        for r in self.attendance:
            if r.get("date") == date_iso and r.get("username") == username:
                return r
        return None

    

    def show_manage_holidays(self):
        """Master-only screen to add/edit/delete yearly holidays."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Manage Holidays")
        dialog.geometry("900x550")
        dialog.configure(bg=COLORS["background"])

        tk.Label(
            dialog,
            text="Manage Yearly Holidays",
            font=("Segoe UI", 16, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(pady=10)

        table_frame = tk.Frame(dialog, bg=COLORS["background"])
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        tree = ttk.Treeview(
            table_frame,
            columns=("Sr. No", "Date", "Day", "Occasion"),
            show="headings",
            height=18,
        )
        for col in ("Sr. No", "Date", "Day", "Occasion"):
            tree.heading(col, text=col)
        tree.column("Sr. No", width=70, anchor=tk.CENTER)
        tree.column("Date", width=140, anchor=tk.CENTER)
        tree.column("Day", width=140, anchor=tk.CENTER)
        tree.column("Occasion", width=420, anchor=tk.W)

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh():
            tree.delete(*tree.get_children())
            # Sort by date if possible
            def sort_key(h):
                d = self._parse_date_input(h.get("date", ""))
                return d or dt_date.max
            self.holidays = sorted(self.holidays, key=sort_key)
            self._reindex_holidays()
            for h in self.holidays:
                tree.insert(
                    "",
                    tk.END,
                    values=(h.get("sr_no", ""), h.get("date", ""), h.get("day", ""), h.get("occasion", "")),
                )

        def open_editor(existing: Optional[Dict] = None):
            ed = tk.Toplevel(dialog)
            ed.title("Edit Holiday" if existing else "Add Holiday")
            ed.geometry("420x320")
            ed.configure(bg=COLORS["surface"])
            ed.transient(dialog)
            #ed.grab_set()

            tk.Label(
                ed,
                text="Edit Holiday" if existing else "Add Holiday",
                font=("Segoe UI", 14, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["primary"],
            ).pack(pady=12)

            form = tk.Frame(ed, bg=COLORS["surface"])
            form.pack(fill=tk.BOTH, expand=True, padx=18, pady=10)

            tk.Label(form, text="Date (YYYY-MM-DD):", bg=COLORS["surface"], fg=COLORS["text"]).pack(anchor=tk.W)
            date_entry = tk.Entry(form, width=32)
            date_entry.pack(pady=(4, 12))

            tk.Label(form, text="Occasion:", bg=COLORS["surface"], fg=COLORS["text"]).pack(anchor=tk.W)
            occ_entry = tk.Entry(form, width=32)
            occ_entry.pack(pady=(4, 12))

            tk.Label(form, text="Day (auto):", bg=COLORS["surface"], fg=COLORS["text"]).pack(anchor=tk.W)
            day_var = tk.StringVar(value="")
            day_entry = tk.Entry(form, width=32, textvariable=day_var, state="readonly")
            day_entry.pack(pady=(4, 12))

            if existing:
                date_entry.insert(0, existing.get("date", ""))
                occ_entry.insert(0, existing.get("occasion", ""))
                day_var.set(existing.get("day", ""))

            def on_date_change(*_):
                d = self._parse_date_input(date_entry.get())
                if d:
                    day_var.set(d.strftime("%A"))
                else:
                    day_var.set("")

            date_entry.bind("<KeyRelease>", on_date_change)
            on_date_change()

            def save():
                d = self._parse_date_input(date_entry.get())
                if not d:
                    messagebox.showerror("Error", "Please enter a valid date (e.g., 2026-01-26).")
                    return
                occasion = occ_entry.get().strip()
                if not occasion:
                    messagebox.showerror("Error", "Occasion cannot be empty.")
                    return

                record = {
                    "sr_no": existing.get("sr_no", "") if existing else "",
                    "date": d.isoformat(),
                    "day": d.strftime("%A"),
                    "occasion": occasion,
                }

                if existing:
                    # Update in-place
                    for i, h in enumerate(self.holidays):
                        if h.get("sr_no") == existing.get("sr_no"):
                            self.holidays[i] = record
                            break
                else:
                    self.holidays.append(record)

                self._reindex_holidays()
                if self.save_data():
                    refresh()
                    ed.destroy()

            btns = tk.Frame(ed, bg=COLORS["surface"])
            btns.pack(pady=10)
            tk.Button(btns, text="Save", bg=COLORS["success"], fg=COLORS["surface"], command=save, padx=18, pady=6).pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text="Cancel", bg=COLORS["text_light"], fg=COLORS["surface"], command=ed.destroy, padx=18, pady=6).pack(side=tk.LEFT, padx=6)

        def add_new():
            open_editor(None)

        def edit_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select", "Please select a holiday to edit.")
                return
            vals = tree.item(sel[0]).get("values", [])
            if not vals:
                return
            sr = str(vals[0])
            existing = next((h for h in self.holidays if h.get("sr_no") == sr), None)
            if existing:
                open_editor(existing)

        def delete_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select", "Please select a holiday to delete.")
                return
            vals = tree.item(sel[0]).get("values", [])
            if not vals:
                return
            sr = str(vals[0])
            if not messagebox.askyesno("Confirm", f"Delete holiday #{sr}?"):
                return
            self.holidays = [h for h in self.holidays if h.get("sr_no") != sr]
            self._reindex_holidays()
            if self.save_data():
                refresh()

        actions = tk.Frame(dialog, bg=COLORS["background"])
        actions.pack(pady=10)
        tk.Button(actions, text="Add", bg=COLORS["success"], fg=COLORS["surface"], command=add_new, padx=18, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(actions, text="Edit", bg=COLORS["secondary"], fg=COLORS["surface"], command=edit_selected, padx=18, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(actions, text="Delete", bg=COLORS["error"], fg=COLORS["surface"], command=delete_selected, padx=18, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(actions, text="Close", bg=COLORS["text_light"], fg=COLORS["surface"], command=dialog.destroy, padx=18, pady=6).pack(side=tk.LEFT, padx=6)

        refresh()






    def show_daily_attendance(self):
    #"""Calendar-based attendance with Department and Employee filters."""

        if self.current_user.get("role") != "master":
            messagebox.showerror(
                "Access Denied",
                "Only the Super User can manage attendance."
            )
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Daily Attendance")
        dialog.geometry("1100x700")
        dialog.configure(bg=COLORS["background"])

        # ---------------------------------------------------------
        # HEADER
        # ---------------------------------------------------------

        header = tk.Frame(
            dialog,
            bg=COLORS["background"]
        )
        header.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(
            header,
            text="Daily Attendance",
            font=("Segoe UI", 18, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(side=tk.LEFT)

        # ---------------------------------------------------------
        # FILTERS
        # ---------------------------------------------------------

        filters = tk.Frame(
            dialog,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        filters.pack(
            fill=tk.X,
            padx=20,
            pady=(0, 10)
        )

        tk.Label(
            filters,
            text="Department:",
            bg=COLORS["surface"],
            fg=COLORS["text"]
        ).pack(side=tk.LEFT, padx=(15, 5), pady=12)

        department_var = tk.StringVar(
            value="All Departments"
        )

        department_values = [
            "All Departments"
        ] + [
            d["department_name"]
            for d in self.departments
            if d.get("is_active", True)
        ]

        department_combo = ttk.Combobox(
            filters,
            textvariable=department_var,
            values=department_values,
            state="readonly",
            width=22
        )

        department_combo.pack(
            side=tk.LEFT,
            padx=(0, 20)
        )

        tk.Label(
            filters,
            text="Employee:",
            bg=COLORS["surface"],
            fg=COLORS["text"]
        ).pack(side=tk.LEFT, padx=(5, 5))

        employee_var = tk.StringVar(
            value="All Employees"
        )

        employee_combo = ttk.Combobox(
            filters,
            textvariable=employee_var,
            state="readonly",
            width=28
        )

        employee_combo.pack(side=tk.LEFT)

        # ---------------------------------------------------------
        # CALENDAR
        # ---------------------------------------------------------

        calendar_frame = tk.Frame(
            dialog,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )

        calendar_frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=10
        )

        month_var = tk.StringVar(
            value=dt_date.today().strftime("%Y-%m")
        )

        # ---------------------------------------------------------
        # CALENDAR FUNCTIONS
        # ---------------------------------------------------------

        def parse_month(value):
            try:
                year, month = map(
                    int,
                    value.split("-")
                )

                if 1 <= month <= 12:
                    return year, month

            except Exception:
                pass

            return None

        def get_filtered_users():
            department = department_var.get()
            employee = employee_var.get()

            users = [
                u for u in self.users.values()
                if u.get("is_active", True)
            ]

            if department != "All Departments":
                users = [
                    u for u in users
                    if u.get("department", "") == department
                ]

            if employee != "All Employees":
                users = [
                    u for u in users
                    if u.get("username") == employee
                ]

            return sorted(
                users,
                key=lambda u: u["username"]
            )

        def refresh_employee_filter(*_):
            department = department_var.get()

            employees = [
                u for u in self.users.values()
                if u.get("is_active", True)
            ]

            if department != "All Departments":
                employees = [
                    u for u in employees
                    if u.get("department", "") == department
                ]

            employee_values = ["All Employees"]

            employee_values += [
                f'{u["username"]} - {u.get("full_name", "")}'
                for u in sorted(
                    employees,
                    key=lambda u: u["username"]
                )
            ]

            employee_combo["values"] = employee_values
            employee_var.set("All Employees")

            refresh_calendar()

        def get_employee_username():
            value = employee_var.get()

            if value == "All Employees":
                return None

            return value.split(" - ")[0]

        def get_status(date_iso, username):
            record = self._get_attendance_record(
                date_iso,
                username
            )

            if record:
                return record.get("status", "")

            return ""

        def show_date_editor(date_value):
            users = get_filtered_users()

            if not users:
                messagebox.showinfo(
                    "Attendance",
                    "No employees found for the selected filters."
                )
                return

            # If All Employees is selected, show employee selection first.
            if len(users) > 1:
                employee_names = [
                    f'{u["username"]} - {u.get("full_name", "")}'
                    for u in users
                ]

                selected = tk.simpledialog.askstring(
                    "Select Employee",
                    "Enter Employee User ID:\n\n"
                    + "\n".join(employee_names)
                )

                if not selected:
                    return

                username = selected.strip()

                if username not in [
                    u["username"] for u in users
                ]:
                    messagebox.showerror(
                        "Error",
                        "Invalid employee User ID."
                    )
                    return

            else:
                username = users[0]["username"]

            user = self.users[username]

            edit = tk.Toplevel(dialog)
            edit.title("Edit Attendance")
            edit.geometry("450x400")
            edit.configure(bg=COLORS["surface"])
            edit.transient(dialog)
            edit.grab_set()

            tk.Label(
                edit,
                text=f"Attendance - {date_value}",
                font=("Segoe UI", 15, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["primary"]
            ).pack(pady=15)

            tk.Label(
                edit,
                text=f'{user["full_name"] or username}',
                font=("Segoe UI", 11),
                bg=COLORS["surface"],
                fg=COLORS["text"]
            ).pack()

            tk.Label(
                edit,
                text=f'Department: {user.get("department", "")}',
                bg=COLORS["surface"],
                fg=COLORS["text_light"]
            ).pack(pady=5)

            tk.Label(
                edit,
                text="Status:",
                bg=COLORS["surface"],
                fg=COLORS["text"]
            ).pack(anchor=tk.W, padx=30, pady=(15, 3))

            status_var = tk.StringVar(
                value=get_status(
                    date_value,
                    username
                ) or "Present"
            )

            status_combo = ttk.Combobox(
                edit,
                textvariable=status_var,
                values=[
                    "Present",
                    "Absent",
                    "Half Day",
                    "Leave",
                    "Holiday",
                    "Work From Home",
                    "Week Off"
                ],
                state="readonly",
                width=28
            )

            status_combo.pack(
                anchor=tk.W,
                padx=30
            )

            tk.Label(
                edit,
                text="Remarks:",
                bg=COLORS["surface"],
                fg=COLORS["text"]
            ).pack(anchor=tk.W, padx=30, pady=(15, 3))

            remarks = tk.Text(
                edit,
                width=40,
                height=5
            )

            remarks.pack(
                padx=30,
                fill=tk.X
            )

            existing = self._get_attendance_record(
                date_value,
                username
            )

            if existing:
                remarks.insert(
                    "1.0",
                    existing.get("remarks", "")
                )

            def save_attendance():
                status = status_var.get()
                remark_text = remarks.get(
                    "1.0",
                    tk.END
                ).strip()

                existing_record = self._get_attendance_record(
                    date_value,
                    username
                )

                now_str = datetime.now().isoformat(
                    sep=" ",
                    timespec="seconds"
                )

                if existing_record:
                    existing_record["status"] = status
                    existing_record["remarks"] = remark_text
                    existing_record["updated_at"] = now_str

                else:
                    self.attendance.append({
                        "date": date_value,
                        "username": username,
                        "status": status,
                        "remarks": remark_text,
                        "updated_at": now_str
                    })

                if self.save_data():
                    edit.destroy()
                    refresh_calendar()

            tk.Button(
                edit,
                text="Save Attendance",
                bg=COLORS["success"],
                fg=COLORS["surface"],
                command=save_attendance,
                padx=20,
                pady=8
            ).pack(pady=20)

        # ---------------------------------------------------------
        # CALENDAR DISPLAY
        # ---------------------------------------------------------

        def refresh_calendar():
            for widget in calendar_frame.winfo_children():
                widget.destroy()

            month_info = parse_month(
                month_var.get()
            )

            if not month_info:
                return

            year, month = month_info

            first_day = dt_date(
                year,
                month,
                1
            )

            if month == 12:
                next_month = dt_date(
                    year + 1,
                    1,
                    1
                )
            else:
                next_month = dt_date(
                    year,
                    month + 1,
                    1
                )

            days_in_month = (
                next_month - first_day
            ).days

            title = tk.Label(
                calendar_frame,
                text=first_day.strftime("%B %Y"),
                font=("Segoe UI", 16, "bold"),
                bg=COLORS["surface"],
                fg=COLORS["primary"]
            )

            title.grid(
                row=0,
                column=0,
                columnspan=7,
                pady=10
            )

            weekdays = [
                "Mon",
                "Tue",
                "Wed",
                "Thu",
                "Fri",
                "Sat",
                "Sun"
            ]

            for col, day_name in enumerate(weekdays):
                tk.Label(
                    calendar_frame,
                    text=day_name,
                    font=("Segoe UI", 10, "bold"),
                    bg=COLORS["background"],
                    fg=COLORS["text"]
                ).grid(
                    row=1,
                    column=col,
                    sticky="nsew",
                    padx=1,
                    pady=1
                )

            selected_employee = get_employee_username()

            for day_number in range(
                1,
                days_in_month + 1
            ):

                current_date = dt_date(
                    year,
                    month,
                    day_number
                )

                row = (
                    (current_date.day
                    + first_day.weekday() - 1)
                    // 7
                ) + 2

                col = current_date.weekday()

                cell = tk.Frame(
                    calendar_frame,
                    bg=COLORS["surface"],
                    highlightbackground=COLORS["border"],
                    highlightthickness=1,
                    width=130,
                    height=80
                )

                cell.grid(
                    row=row,
                    column=col,
                    sticky="nsew",
                    padx=2,
                    pady=2
                )

                cell.grid_propagate(False)

                date_iso = current_date.isoformat()

                tk.Label(
                    cell,
                    text=str(day_number),
                    font=("Segoe UI", 10, "bold"),
                    bg=COLORS["surface"],
                    fg=COLORS["primary"]
                ).pack(
                    anchor=tk.NW,
                    padx=5,
                    pady=3
                )

                if selected_employee:

                    status = get_status(
                        date_iso,
                        selected_employee
                    )

                    tk.Label(
                        cell,
                        text=status or "-",
                        font=("Segoe UI", 9),
                        bg=COLORS["surface"],
                        fg=COLORS["text"]
                    ).pack(
                        pady=5
                    )

                else:

                    users = get_filtered_users()

                    present = 0
                    absent = 0
                    leave = 0
                    wfh = 0

                    for user in users:

                        status = get_status(
                            date_iso,
                            user["username"]
                        )

                        if status == "Present":
                            present += 1

                        elif status == "Absent":
                            absent += 1

                        elif status == "Leave":
                            leave += 1

                        elif status == "Work From Home":
                            wfh += 1

                    tk.Label(
                        cell,
                        text=(
                            f"P:{present} "
                            f"A:{absent} "
                            f"L:{leave} "
                            f"W:{wfh}"
                        ),
                        font=("Segoe UI", 8),
                        bg=COLORS["surface"],
                        fg=COLORS["text"]
                    ).pack(
                        pady=5
                    )

                # Click date
                cell.bind(
                    "<Button-1>",
                    lambda e, d=date_iso:
                        show_date_editor(d)
                )

                for child in cell.winfo_children():
                    child.bind(
                        "<Button-1>",
                        lambda e, d=date_iso:
                            show_date_editor(d)
                    )

            for i in range(7):
                calendar_frame.grid_columnconfigure(
                    i,
                    weight=1
                )

        # ---------------------------------------------------------
        # NAVIGATION
        # ---------------------------------------------------------

        nav = tk.Frame(
            dialog,
            bg=COLORS["background"]
        )

        nav.pack(
            fill=tk.X,
            padx=20,
            pady=10
        )

        def previous_month():
            info = parse_month(
                month_var.get()
            )

            if not info:
                return

            year, month = info

            if month == 1:
                year -= 1
                month = 12
            else:
                month -= 1

            month_var.set(
                f"{year}-{month:02d}"
            )

            refresh_calendar()

        def next_month():
            info = parse_month(
                month_var.get()
            )

            if not info:
                return

            year, month = info

            if month == 12:
                year += 1
                month = 1
            else:
                month += 1

            month_var.set(
                f"{year}-{month:02d}"
            )

            refresh_calendar()

        tk.Button(
            nav,
            text="< Previous",
            command=previous_month,
            padx=15
        ).pack(side=tk.LEFT)

        tk.Button(
            nav,
            text="Next >",
            command=next_month,
            padx=15
        ).pack(side=tk.LEFT, padx=5)

        def refresh_attendance_from_excel():
            """Reload attendance from Excel."""

            try:

                # Reload users in case users were changed
                self.users = load_users(
                    self.users_ws
                )

                # Reload attendance from Excel
                self.attendance = load_attendance(
                    self.attendance_ws,
                    self.users
                )

                # Redraw calendar
                refresh_calendar()

            except Exception as e:

                messagebox.showerror(
                    "Refresh Error",
                    f"Could not refresh attendance data:\n\n{e}"
                )


        tk.Button(
            nav,
            text="🔄 Refresh",
            command=refresh_attendance_from_excel,
            padx=15
        ).pack(
            side=tk.LEFT,
            padx=5
        )


        tk.Button(
            nav,
            text="Close",
            command=dialog.destroy,
            padx=15
        ).pack(side=tk.RIGHT)


        tk.Button(
            nav,
            text="Close",
            command=dialog.destroy,
            padx=15
        ).pack(side=tk.RIGHT)

        department_combo.bind(
            "<<ComboboxSelected>>",
            refresh_employee_filter
        )

        employee_combo.bind(
            "<<ComboboxSelected>>",
            lambda e: refresh_calendar()
        )

        refresh_employee_filter()
    





    def show_attendance_summary(self):
        """Matrix-style monthly attendance summary (like your screenshot)."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Attendance Summary")
        dialog.geometry("1100x550")
        dialog.configure(bg=COLORS["background"])

        header = tk.Frame(dialog, bg=COLORS["background"])
        header.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(
            header,
            text="Attendance Summary (Matrix View)",
            font=("Segoe UI", 16, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(side=tk.LEFT)

        tk.Label(
            header,
            text="Month (YYYY-MM):",
            font=("Segoe UI", 10),
            bg=COLORS["background"],
            fg=COLORS["text"],
        ).pack(side=tk.LEFT, padx=(40, 6))

        month_var = tk.StringVar(
            value=f"{dt_date.today().year}-{dt_date.today().month:02d}"
        )
        month_entry = tk.Entry(header, textvariable=month_var, width=8)
        month_entry.pack(side=tk.LEFT)

        content = tk.Frame(dialog, bg=COLORS["background"])
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Build Treeview with dynamic date columns
        tree = ttk.Treeview(content, show="headings", height=18)

        xscroll = ttk.Scrollbar(content, orient=tk.HORIZONTAL, command=tree.xview)
        yscroll = ttk.Scrollbar(content, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.LEFT, fill=tk.Y)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)

        def parse_month(text: str) -> Optional[Tuple[int, int]]:
            s = (text or "").strip()
            try:
                parts = s.split("-")
                if len(parts) != 2:
                    return None
                year = int(parts[0])
                month = int(parts[1])
                if 1 <= month <= 12:
                    return year, month
                return None
            except Exception:
                return None

        def build_dates(year: int, month: int) -> List[dt_date]:
            first = dt_date(year, month, 1)
            # Next month
            if month == 12:
                nxt = dt_date(year + 1, 1, 1)
            else:
                nxt = dt_date(year, month + 1, 1)
            days = []
            cur = first
            while cur < nxt:
                days.append(cur)
                cur = cur.replace(day=cur.day + 1)
            return days

        def refresh():
            month_info = parse_month(month_var.get())
            if not month_info:
                messagebox.showerror(
                    "Invalid month", "Please enter month as YYYY-MM, e.g. 2026-02."
                )
                return
            year, month = month_info
            days = build_dates(year, month)

            # Configure columns: Employee Name + one per day
            cols = ["Employee Name"] + [d.strftime("%d-%b-%y") for d in days]
            tree["columns"] = cols
            for c in cols:
                tree.heading(c, text=c)
                width = 150 if c == "Employee Name" else 80
                anchor = tk.W if c == "Employee Name" else tk.CENTER
                tree.column(c, width=width, anchor=anchor)

            tree.delete(*tree.get_children())

            active_users = [
                u for u in self.users.values() if u.get("is_active", True)
            ]
            active_users = sorted(active_users, key=lambda u: u["username"])

            # Pre-index attendance by (date, username)
            index: Dict[Tuple[str, str], str] = {}
            for r in self.attendance:
                key = (r.get("date", ""), r.get("username", ""))
                index[key] = r.get("status", "") or ""

            for u in active_users:
                row_vals = [u.get("full_name") or u["username"]]
                for d in days:
                    key = (d.isoformat(), u["username"])
                    val = index.get(key, "")
                    row_vals.append(val or "")
                tree.insert("", tk.END, values=row_vals)

        footer = tk.Frame(dialog, bg=COLORS["background"])
        footer.pack(fill=tk.X, padx=20, pady=10)

        tk.Button(
            footer,
            text="Load Month",
            bg=COLORS["secondary"],
            fg=COLORS["surface"],
            command=refresh,
            padx=18,
            pady=6,
        ).pack(side=tk.LEFT)

        tk.Button(
            footer,
            text="Close",
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            command=dialog.destroy,
            padx=18,
            pady=6,
        ).pack(side=tk.RIGHT)

        refresh()

    def show_my_attendance(self):
        """
        Show attendance based on department access.

        Normal user:
            - Can see attendance of employees in their own department only.
            - Cannot see employees from other departments.

        Master user:
            - Can see attendance across all departments.
        """

        # =========================================================
        # SECURITY CHECK
        # =========================================================

        if not self.current_user:
            messagebox.showerror(
                "Access Denied",
                "You must be logged in."
            )
            return

        current_role = str(
            self.current_user.get("role", "")
        ).strip().lower()

        current_username = str(
            self.current_user.get("username", "")
        ).strip()

        current_department = str(
            self.current_user.get("department", "")
        ).strip()

        # =========================================================
        # DETERMINE WHICH EMPLOYEES CAN BE VIEWED
        # =========================================================

        if current_role == "master":

            # Master can see everybody.
            visible_users = {
                str(user.get("username", "")).strip(): user
                for user in self.users.values()
                if user.get("is_active", True)
            }

            window_title = "All Employee Attendance"
            header_text = "Master User - All Departments"

        else:

            # Normal users can only see employees
            # belonging to their own department.
            visible_users = {
                str(user.get("username", "")).strip(): user
                for user in self.users.values()
                if (
                    user.get("is_active", True)
                    and str(
                        user.get("department", "")
                    ).strip().lower()
                    == current_department.lower()
                )
            }

            window_title = (
                f"{current_department} Department Attendance"
            )

            header_text = (
                f"Department: {current_department}"
            )

        # =========================================================
        # CREATE WINDOW
        # =========================================================

        dialog = tk.Toplevel(self.root)

        dialog.title(
            window_title
        )

        dialog.geometry(
            "1100x650"
        )

        dialog.configure(
            bg=COLORS["background"]
        )

        dialog.transient(
            self.root
        )

        # =========================================================
        # HEADER
        # =========================================================

        tk.Label(
            dialog,
            text="Attendance",
            font=("Segoe UI", 18, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(
            pady=(20, 5)
        )

        tk.Label(
            dialog,
            text=header_text,
            font=("Segoe UI", 11),
            bg=COLORS["background"],
            fg=COLORS["text"]
        ).pack(
            pady=(0, 15)
        )

        # =========================================================
        # TABLE
        # =========================================================

        table_frame = tk.Frame(
            dialog,
            bg=COLORS["surface"]
        )

        table_frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=10
        )

        columns = (
            "Employee",
            "Department",
            "Date",
            "Day",
            "Status",
            "Remarks"
        )

        tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings"
        )

        # =========================================================
        # HEADINGS
        # =========================================================

        for column in columns:

            tree.heading(
                column,
                text=column
            )

        # =========================================================
        # COLUMN WIDTHS
        # =========================================================

        tree.column(
            "Employee",
            width=190,
            anchor=tk.W
        )

        tree.column(
            "Department",
            width=150,
            anchor=tk.W
        )

        tree.column(
            "Date",
            width=120,
            anchor=tk.CENTER
        )

        tree.column(
            "Day",
            width=120,
            anchor=tk.CENTER
        )

        tree.column(
            "Status",
            width=180,
            anchor=tk.CENTER
        )

        tree.column(
            "Remarks",
            width=300,
            anchor=tk.W
        )

        # =========================================================
        # SCROLLBARS
        # =========================================================

        vertical_scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.VERTICAL,
            command=tree.yview
        )

        horizontal_scrollbar = ttk.Scrollbar(
            table_frame,
            orient=tk.HORIZONTAL,
            command=tree.xview
        )

        tree.configure(
            yscrollcommand=vertical_scrollbar.set,
            xscrollcommand=horizontal_scrollbar.set
        )

        tree.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True
        )

        vertical_scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y
        )

        horizontal_scrollbar.pack(
            side=tk.BOTTOM,
            fill=tk.X
        )

        # =========================================================
        # FILTER ATTENDANCE BY DEPARTMENT
        # =========================================================

        visible_records = []

        for record in self.attendance:

            username = str(
                record.get("username", "")
            ).strip()

            # -----------------------------------------------------
            # Employee must exist in Users
            # -----------------------------------------------------

            employee = visible_users.get(
                username
            )

            if not employee:
                continue

            # -----------------------------------------------------
            # Extra protection for normal users
            # -----------------------------------------------------

            if current_role != "master":

                employee_department = str(
                    employee.get("department", "")
                ).strip()

                if (
                    employee_department.lower()
                    != current_department.lower()
                ):
                    continue

            visible_records.append(
                record
            )

        # =========================================================
        # SORT
        # =========================================================

        visible_records.sort(
            key=lambda record: (
                str(
                    record.get("date", "")
                ),
                str(
                    record.get("employee_name", "")
                ).lower()
            ),
            reverse=True
        )

        # =========================================================
        # DISPLAY
        # =========================================================

        for record in visible_records:

            username = str(
                record.get("username", "")
            ).strip()

            employee = visible_users.get(
                username,
                {}
            )

            employee_name = str(
                employee.get(
                    "full_name",
                    record.get(
                        "employee_name",
                        username
                    )
                )
            ).strip()

            department = str(
                employee.get(
                    "department",
                    ""
                )
            ).strip()

            date_value = str(
                record.get(
                    "date",
                    ""
                )
            ).strip()

            try:

                date_obj = datetime.strptime(
                    date_value,
                    "%Y-%m-%d"
                )

                display_date = date_obj.strftime(
                    "%d-%b-%Y"
                )

                day_name = date_obj.strftime(
                    "%A"
                )

            except Exception:

                display_date = date_value
                day_name = ""

            tree.insert(
                "",
                tk.END,
                values=(
                    employee_name,
                    department,
                    display_date,
                    day_name,
                    record.get(
                        "status",
                        ""
                    ),
                    record.get(
                        "remarks",
                        ""
                    )
                )
            )

        # =========================================================
        # NO RECORDS
        # =========================================================

        if not visible_records:

            tk.Label(
                dialog,
                text=(
                    "No attendance records found "
                    f"for the {current_department} department."
                    if current_role != "master"
                    else "No attendance records found."
                ),
                font=("Segoe UI", 11),
                bg=COLORS["background"],
                fg=COLORS["text_light"]
            ).pack(
                pady=10
            )

        # =========================================================
        # CLOSE
        # =========================================================

        tk.Button(
            dialog,
            text="Close",
            font=("Segoe UI", 11),
            bg=COLORS["text_light"],
            fg=COLORS["surface"],
            command=dialog.destroy,
            padx=25,
            pady=8
        ).pack(
            pady=15
        )
            
    def show_user_dashboard(self):
        """Display a professional employee dashboard."""
        self.clear_main_frame()

        # =====================================================
        # Dashboard header
        # =====================================================

        header = tk.Frame(
            self.main_frame,
            bg=COLORS["background"]
        )

        header.pack(
            fill=tk.X,
            padx=25,
            pady=(5, 15)
        )

        # =====================================================
        # Department Employees
        # =====================================================

        self.show_department_employees_box(
            self.main_frame
        )

        # -----------------------------------------------------
        # Dashboard title
        # -----------------------------------------------------

        tk.Label(
            header,
            text="Master Dashboard",
            font=("Segoe UI", 22, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"]
        ).pack(
            side=tk.LEFT
        )

        # -----------------------------------------------------
        # Logged-in user information
        # -----------------------------------------------------

        user_info_frame = tk.Frame(
            header,
            bg=COLORS["background"]
        )

        user_info_frame.pack(
            side=tk.RIGHT,
            padx=10
        )

        # Name
        tk.Label(
            user_info_frame,
            text=(
                f"Welcome, "
                f"{self.current_user.get('full_name') or self.current_user['username']}"
            ),
            font=("Segoe UI", 11, "bold"),
            bg=COLORS["background"],
            fg=COLORS["primary"],
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )

        # Designation
        tk.Label(
            user_info_frame,
            text=(
                f"Designation: "
                f"{self.current_user.get('designation') or 'N/A'}"
            ),
            font=("Segoe UI", 9),
            bg=COLORS["background"],
            fg=COLORS["text_light"],
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )

        # Status
        employee_status = (
            "Active"
            if self.current_user.get("is_active", True)
            else "Inactive"
        )

        tk.Label(
            user_info_frame,
            text=f"Status: {employee_status}",
            font=("Segoe UI", 9, "bold"),
            bg=COLORS["background"],
            fg=(
                COLORS["success"]
                if employee_status == "Active"
                else COLORS["error"]
            ),
            anchor=tk.E
        ).pack(
            anchor=tk.E
        )


        # =====================================================
        # Two-column dashboard
        # =====================================================
        content = tk.Frame(
            self.main_frame,
            bg=COLORS["background"]
        )
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 10))
        content.columnconfigure(0, weight=0, minsize=315)
        content.columnconfigure(1, weight=1)
        content.rowconfigure(0, weight=1)

        # =====================================================
        # Left: Employee menu
        # =====================================================
        menu_card = tk.Frame(
            content,
            bg=COLORS["surface"],
            highlightbackground=COLORS["border"],
            highlightthickness=1
        )
        menu_card.grid(row=0, column=0, sticky="nsew", padx=(0, 12))

        # =====================================================
        # My Workspace
        # =====================================================

        tk.Label(
            menu_card,
            text="My Workspace",
            font=("Segoe UI", 14, "bold"),
            bg=COLORS["surface"],
            fg=COLORS["primary"]
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(18, 3)
        )

        tk.Label(
            menu_card,
            text="Access your profile, leave and attendance tools",
            font=("Segoe UI", 9),
            bg=COLORS["surface"],
            fg=COLORS["text_light"],
            wraplength=270,
            justify=tk.LEFT
        ).pack(
            anchor=tk.W,
            padx=18,
            pady=(0, 10)
        )

        # =====================================================
        # SCROLLABLE WORKSPACE
        # =====================================================

        workspace_container = tk.Frame(
            menu_card,
            bg=COLORS["surface"]
        )

        workspace_container.pack(
            fill=tk.BOTH,
            expand=True,
            padx=8,
            pady=(0, 8)
        )

        # -----------------------------------------------------
        # Canvas
        # -----------------------------------------------------

        workspace_canvas = tk.Canvas(
            workspace_container,
            bg=COLORS["surface"],
            highlightthickness=0,
            bd=0
        )

        workspace_canvas.pack(
            side=tk.LEFT,
            fill=tk.BOTH,
            expand=True
        )

        # -----------------------------------------------------
        # Scrollbar
        # -----------------------------------------------------

        workspace_scrollbar = ttk.Scrollbar(
            workspace_container,
            orient=tk.VERTICAL,
            command=workspace_canvas.yview
        )

        workspace_scrollbar.pack(
            side=tk.RIGHT,
            fill=tk.Y
        )

        workspace_canvas.configure(
            yscrollcommand=workspace_scrollbar.set
        )

        # -----------------------------------------------------
        # Inner frame
        # -----------------------------------------------------

        workspace_buttons_frame = tk.Frame(
            workspace_canvas,
            bg=COLORS["surface"]
        )

        workspace_window = workspace_canvas.create_window(
            (0, 0),
            window=workspace_buttons_frame,
            anchor="nw"
        )

        # =====================================================
        # UPDATE SCROLL REGION
        # =====================================================

        def update_workspace_scroll_region(event=None):
            workspace_canvas.configure(
                scrollregion=workspace_canvas.bbox("all")
            )

        workspace_buttons_frame.bind(
            "<Configure>",
            update_workspace_scroll_region
        )

        # =====================================================
        # KEEP INNER FRAME SAME WIDTH AS CANVAS
        # =====================================================

        def update_workspace_width(event):
            workspace_canvas.itemconfigure(
                workspace_window,
                width=event.width
            )

        workspace_canvas.bind(
            "<Configure>",
            update_workspace_width
        )

        # =====================================================
        # MOUSE WHEEL
        # =====================================================

        def workspace_mousewheel(event):
            workspace_canvas.yview_scroll(
                int(-1 * (event.delta / 120)),
                "units"
            )

        workspace_canvas.bind(
            "<Enter>",
            lambda event: workspace_canvas.bind_all(
                "<MouseWheel>",
                workspace_mousewheel
            )
        )

        workspace_canvas.bind(
            "<Leave>",
            lambda event: workspace_canvas.unbind_all(
                "<MouseWheel>"
            )
        )

        # =====================================================
        # WORKSPACE BUTTONS
        # =====================================================

        buttons = [
            ("My Profile", self.show_my_profile),
            ("Apply for Leave", self.show_apply_leave_dialog),
            ("My Leaves", self.show_my_leaves),
            ("My Attendance", self.show_my_attendance),
            ("Yearly Leaves", self.show_yearly_leaves),
            ("Cancel Leave", self.show_cancel_leave_dialog),
            ("💬 Messages", self.open_messenger),
        ]

        for text, command in buttons:

            btn = tk.Button(
                workspace_buttons_frame,
                text=text,
                font=("Segoe UI", 10, "bold"),
                bg=COLORS["secondary"],
                fg=COLORS["surface"],
                activebackground=COLORS["accent"],
                activeforeground=COLORS["surface"],
                relief=tk.FLAT,
                cursor="hand2",
                anchor=tk.W,
                padx=16,
                pady=9,
                command=command
            )

            btn.pack(
                fill=tk.X,
                padx=8,
                pady=4
            )

        # =====================================================
        # LOGOUT
        # =====================================================

        tk.Frame(
            workspace_buttons_frame,
            bg=COLORS["border"],
            height=1
        ).pack(
            fill=tk.X,
            padx=8,
            pady=(12, 10)
        )

        tk.Button(
            workspace_buttons_frame,
            text="Logout",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["error"],
            fg=COLORS["surface"],
            activebackground="#c0392b",
            activeforeground=COLORS["surface"],
            relief=tk.FLAT,
            cursor="hand2",
            anchor=tk.W,
            padx=16,
            pady=9,
            command=self.logout
        ).pack(
            fill=tk.X,
            padx=8,
            pady=(0, 12)
        )

        # =====================================================
        # Right: Announcements
        # =====================================================
        self.show_announcements_section(
            content,
            master_controls=False,
            grid_position=(0, 1)
        )

    def show_add_user_dialog(self):
        """Show dialog to add a new user."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add User")
        dialog.geometry("500x650")
        dialog.configure(bg=COLORS['surface'])
        dialog.transient(self.root)
        #dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (650 // 2)
        dialog.geometry(f"500x650+{x}+{y}")
        
        # Main container frame
        main_container = tk.Frame(dialog, bg=COLORS['surface'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Title
        title_label = tk.Label(
            main_container,
            text="Add New User",
            font=('Segoe UI', 18, 'bold'),
            bg=COLORS['surface'],
            fg=COLORS['primary']
        )
        title_label.pack(pady=(0, 25))
        
        # Form fields container
        form_frame = tk.Frame(main_container, bg=COLORS['surface'])
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Username
        tk.Label(form_frame, text="User ID:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(anchor=tk.W, pady=(0, 5))
        username_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=35, relief=tk.SOLID, bd=1)
        username_entry.pack(pady=(0, 15))
        
        # Full name
        tk.Label(form_frame, text="Full Name (optional):", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(anchor=tk.W, pady=(0, 5))
        fullname_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=35, relief=tk.SOLID, bd=1)
        fullname_entry.pack(pady=(0, 15))
        
        # Role
        tk.Label(form_frame, text="Role:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(anchor=tk.W, pady=(0, 5))
        role_var = tk.StringVar(value="user")
        role_frame = tk.Frame(form_frame, bg=COLORS['surface'])
        role_frame.pack(anchor=tk.W, pady=(0, 15))
        tk.Radiobutton(
            role_frame,
            text="User",
            variable=role_var,
            value="user",
            bg=COLORS['surface'],
            font=('Segoe UI', 10),
            activebackground=COLORS['surface']
        ).pack(side=tk.LEFT, padx=(0, 20))
        tk.Radiobutton(
            role_frame,
            text="Master",
            variable=role_var,
            value="master",
            bg=COLORS['surface'],
            font=('Segoe UI', 10),
            activebackground=COLORS['surface']
        ).pack(side=tk.LEFT)

        # Department
        tk.Label(
            form_frame,
            text="Department:",
            font=('Segoe UI', 11),
            bg=COLORS['surface'],
            fg=COLORS['text']
        ).pack(anchor=tk.W, pady=(0, 5))

        department_var = tk.StringVar()

        active_departments = [
            d["department_name"]
            for d in self.departments
            if d.get("is_active", True)
        ]

        department_combo = ttk.Combobox(
            form_frame,
            textvariable=department_var,
            values=active_departments,
            state="readonly",
            width=32
        )

        department_combo.pack(pady=(0, 15))
        
        # Password
        tk.Label(form_frame, text="Password:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(anchor=tk.W, pady=(0, 5))
        password_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=35, show="*", relief=tk.SOLID, bd=1)
        password_entry.pack(pady=(0, 15))
        
        # Confirm password
        tk.Label(form_frame, text="Confirm Password:", font=('Segoe UI', 11), bg=COLORS['surface'], fg=COLORS['text']).pack(anchor=tk.W, pady=(0, 5))
        confirm_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=35, show="*", relief=tk.SOLID, bd=1)
        confirm_entry.pack(pady=(0, 25))
        
        def add_user():
            username = username_entry.get().strip()
            department = department_var.get().strip()
            full_name = fullname_entry.get().strip()
            role = role_var.get()
            password = password_entry.get()
            confirm = confirm_entry.get()

            if not department:
                messagebox.showerror("Error", "Please select a department.")
                return

            if not username:
                messagebox.showerror("Error", "User ID cannot be empty.")
                return

            if username in self.users:
                messagebox.showerror("Error", "User ID already exists.")
                return

            if password != confirm:
                messagebox.showerror("Error", "Passwords do not match.")
                return

            if not password:
                messagebox.showerror("Error", "Password cannot be empty.")
                return

            self.users[username] = {
                "username": username,
                "full_name": full_name,
                "department": department,
                "password_hash": hash_password(password),
                "role": role,
                "is_active": True,
            }

            if self.save_data():
                messagebox.showinfo(
                    "Success",
                    f"User '{username}' created successfully."
                )
                dialog.destroy()
                
            # Button frame
        button_frame = tk.Frame(
            main_container,
            bg=COLORS['surface']
        )
        button_frame.pack(pady=(10, 0))

        # Add User button
        add_btn = tk.Button(
            button_frame,
            text="Add User",
            font=('Segoe UI', 12, 'bold'),
            bg=COLORS['success'],
            fg=COLORS['surface'],
            activebackground='#229954',
            activeforeground=COLORS['surface'],
            relief=tk.FLAT,
            cursor='hand2',
            command=add_user,
            padx=30,
            pady=10,
            width=15
        )
        add_btn.pack()

        # Cancel button
        cancel_btn = tk.Button(
            button_frame,
            text="Cancel",
            font=('Segoe UI', 11),
            bg=COLORS['text_light'],
            fg=COLORS['surface'],
            activebackground='#5d6d7e',
            activeforeground=COLORS['surface'],
            relief=tk.FLAT,
            cursor='hand2',
            command=dialog.destroy,
            padx=20,
            pady=8,
            width=12
        )
        cancel_btn.pack(pady=(10, 0))

        username_entry.focus()

        
    def show_manage_users(self):
        """Show user management window."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Manage Users")
        dialog.geometry("700x500")
        dialog.configure(bg=COLORS['background'])
        
        tk.Label(dialog, text="User Management", font=('Segoe UI', 16, 'bold'), bg=COLORS['background'], fg=COLORS['primary']).pack(pady=10)
        
        # Treeview for users
        tree_frame = tk.Frame(dialog, bg=COLORS['background'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=("Username", "Full Name", "Role", "Status"), show="headings", height=15)
        tree.heading("Username", text="Username")
        tree.heading("Full Name", text="Full Name")
        tree.heading("Role", text="Role")
        tree.heading("Status", text="Status")
        tree.column("Username", width=150)
        tree.column("Full Name", width=200)
        tree.column("Role", width=100)
        tree.column("Status", width=100)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def delete_user():
            selected = tree.selection()

            if not selected:
                messagebox.showwarning(
                    "No Selection",
                    "Please select a user to delete."
                )
                return

            item = tree.item(selected[0])
            values = item.get("values", [])

            if not values:
                return

            username = str(values[0]).strip()

            if username not in self.users:
                messagebox.showerror(
                    "Error",
                    f"User '{username}' was not found."
                )
                return

            # Prevent deleting the currently logged-in account
            if username == self.current_user:
                messagebox.showerror(
                    "Cannot Delete",
                    "You cannot delete the account you are currently logged in with."
                )
                return

            user = self.users[username]

            employee_name = user.get(
                "full_name",
                username
            )

            # First confirmation
            confirm = messagebox.askyesno(
                "Delete User",
                f"Are you sure you want to permanently delete:\n\n"
                f"Username: {username}\n"
                f"Name: {employee_name}\n\n"
                f"This action cannot be undone."
            )

            if not confirm:
                return

            # Permanent deletion
            del self.users[username]

            if self.save_data():

                messagebox.showinfo(
                    "User Deleted",
                    f"User '{username}' has been permanently deleted."
                )

                refresh_users()

        def reactivate_user():
            selected = tree.selection()

            if not selected:
                messagebox.showwarning(
                    "No Selection",
                    "Please select a user to reactivate."
                )
                return

            item = tree.item(selected[0])

            # Change this index if your Treeview columns differ
            username = item["values"][0]

            if username not in self.users:
                messagebox.showerror(
                    "Error",
                    f"User '{username}' was not found."
                )
                return

            user = self.users[username]

            if user.get("is_active", True):
                messagebox.showinfo(
                    "Already Active",
                    f"User '{username}' is already active."
                )
                return

            confirm = messagebox.askyesno(
                "Reactivate User",
                f"Are you sure you want to reactivate '{username}'?"
            )

            if not confirm:
                return

            user["is_active"] = True

            if self.save_data():
                messagebox.showinfo(
                    "Success",
                    f"User '{username}' has been reactivated."
                )

                refresh_users()
        


        # Populate tree
        for user in self.users.values():
            status = "Active" if user.get("is_active", True) else "Inactive"
            tree.insert("", tk.END, values=(
                user["username"],
                user.get("full_name", ""),
                user.get("role", "user"),
                status
            ))
            
        def deactivate_user():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a user.")
                return
            item = tree.item(selected[0])
            username = item['values'][0]
            if username == self.current_user["username"]:
                messagebox.showerror("Error", "You cannot deactivate yourself.")
                return
            user = self.users.get(username)
            if user:
                user["is_active"] = False
                if self.save_data():
                    messagebox.showinfo("Success", f"User '{username}' deactivated.")
                    dialog.destroy()
                    self.show_master_dashboard()
                    
        tk.Button(
            dialog,
            text="Deactivate Selected User",
            font=('Segoe UI', 10),
            bg=COLORS['error'],
            fg=COLORS['surface'],
            command=deactivate_user,
            padx=15,
            pady=5
        ).pack(pady=10)

        tk.Button(
            dialog,
            text="Reactivate User",
            font=('Segoe UI', 10),
            bg=COLORS['error'],
            fg=COLORS['surface'],
            command=reactivate_user,
            padx=15,
            pady=5
        ).pack(pady=10)

        tk.Button(
                dialog,
                text="Delete User",
                font=('Segoe UI', 10),
                bg=COLORS['error'],
                fg=COLORS['surface'],
                command=delete_user,
                padx=15,
                pady=5
            ).pack(pady=10)

        
                
    def show_approve_leaves(self):
        """Show pending leaves for approval."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Approve Leaves")
        dialog.geometry("900x500")
        dialog.configure(bg=COLORS['background'])
        
        tk.Label(dialog, text="Pending Leave Requests", font=('Segoe UI', 16, 'bold'), bg=COLORS['background'], fg=COLORS['primary']).pack(pady=10)
        
        pending = [lv for lv in self.leaves if lv["status"] == "PENDING"]
        if not pending:
            tk.Label(dialog, text="No pending leave requests.", font=('Segoe UI', 12), bg=COLORS['background'], fg=COLORS['text_light']).pack(pady=50)
            return
            
        tree_frame = tk.Frame(dialog, bg=COLORS['background'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=("ID", "User", "Start Date", "End Date", "Type", "Reason"), show="headings", height=12)
        tree.heading("ID", text="ID")
        tree.heading("User", text="User")
        tree.heading("Start Date", text="Start Date")
        tree.heading("End Date", text="End Date")
        tree.heading("Type", text="Type")
        tree.heading("Reason", text="Reason")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for lv in pending:
            tree.insert("", tk.END, values=(
                lv["leave_id"],
                lv["username"],
                lv["start_date"],
                lv["end_date"],
                lv["leave_type"],
                lv["reason"][:50] + "..." if len(lv["reason"]) > 50 else lv["reason"]
            ))
            
        def approve_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a leave request.")
                return
            item = tree.item(selected[0])
            leave_id = int(item['values'][0])
            target = next((lv for lv in self.leaves if lv["leave_id"] == leave_id), None)
            if target:
                target["status"] = "APPROVED"
                target["updated_at"] = datetime.now().isoformat(sep=" ", timespec="seconds")
                if self.save_data():
                    messagebox.showinfo("Success", f"Leave {leave_id} approved.")
                    dialog.destroy()
                    self.show_master_dashboard()
                    
        def reject_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a leave request.")
                return
            item = tree.item(selected[0])
            leave_id = int(item['values'][0])
            target = next((lv for lv in self.leaves if lv["leave_id"] == leave_id), None)
            if target:
                target["status"] = "REJECTED"
                target["updated_at"] = datetime.now().isoformat(sep=" ", timespec="seconds")
                if self.save_data():
                    messagebox.showinfo("Success", f"Leave {leave_id} rejected.")
                    dialog.destroy()
                    self.show_master_dashboard()
                    
        btn_frame = tk.Frame(dialog, bg=COLORS['background'])
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="Approve", font=('Segoe UI', 10), bg=COLORS['success'], fg=COLORS['surface'], command=approve_selected, padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Reject", font=('Segoe UI', 10), bg=COLORS['error'], fg=COLORS['surface'], command=reject_selected, padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        
    def show_all_leaves(self):
        """Show all leave requests."""
        dialog = tk.Toplevel(self.root)
        dialog.title("All Leaves")
        dialog.geometry("1000x600")
        dialog.configure(bg=COLORS['background'])
        
        tk.Label(dialog, text="All Leave Requests", font=('Segoe UI', 16, 'bold'), bg=COLORS['background'], fg=COLORS['primary']).pack(pady=10)
        
        tree_frame = tk.Frame(dialog, bg=COLORS['background'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=("ID", "User", "Start Date", "End Date", "Type", "Status", "Reason"), show="headings", height=20)
        for col in ("ID", "User", "Start Date", "End Date", "Type", "Status", "Reason"):
            tree.heading(col, text=col)
            tree.column(col, width=120)
        tree.column("Reason", width=200)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for lv in self.leaves:
            tree.insert("", tk.END, values=(
                lv["leave_id"],
                lv["username"],
                lv["start_date"],
                lv["end_date"],
                lv["leave_type"],
                lv["status"],
                lv["reason"][:50] + "..." if len(lv["reason"]) > 50 else lv["reason"]
            ))
            
    def show_apply_leave_dialog(self):
        """Show dialog to apply for leave."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Apply for Leave")
        dialog.geometry("450x400")
        dialog.configure(bg=COLORS['surface'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"450x400+{x}+{y}")
        
        tk.Label(dialog, text="Apply for Leave", font=('Segoe UI', 16, 'bold'), bg=COLORS['surface'], fg=COLORS['primary']).pack(pady=20)
        
        tk.Label(dialog, text="Start Date (YYYY-MM-DD):", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        start_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30)
        start_entry.pack(pady=5)
        
        tk.Label(dialog, text="End Date (YYYY-MM-DD):", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        end_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30)
        end_entry.pack(pady=5)
        
        tk.Label(dialog, text="Leave Type:", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        type_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30)
        type_entry.pack(pady=5)
        type_entry.insert(0, "General")
        
        tk.Label(dialog, text="Reason:", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        reason_text = tk.Text(dialog, font=('Segoe UI', 10), width=30, height=4)
        reason_text.pack(pady=5)
        
        def submit_leave():
            start_date = start_entry.get().strip()
            end_date = end_entry.get().strip()
            leave_type = type_entry.get().strip() or "General"
            reason = reason_text.get("1.0", tk.END).strip()
            
            try:
                sdt = datetime.strptime(start_date, "%Y-%m-%d").date()
                edt = datetime.strptime(end_date, "%Y-%m-%d").date()
                if edt < sdt:
                    messagebox.showerror("Error", "End date cannot be before start date.")
                    return
            except ValueError:
                messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD.")
                return
                
            leave_id = generate_next_leave_id(self.leaves)
            now = datetime.now().isoformat(sep=" ", timespec="seconds")
            self.leaves.append({
                "leave_id": leave_id,
                "username": self.current_user["username"],
                "start_date": start_date,
                "end_date": end_date,
                "leave_type": leave_type,
                "status": "PENDING",
                "reason": reason,
                "created_at": now,
                "updated_at": now,
            })
            if self.save_data():
                messagebox.showinfo("Success", f"Leave request created with ID {leave_id}.")
                dialog.destroy()
                
        tk.Button(
            dialog,
            text="Submit",
            font=('Segoe UI', 11, 'bold'),
            bg=COLORS['success'],
            fg=COLORS['surface'],
            command=submit_leave,
            padx=20,
            pady=5
        ).pack(pady=20)
        
    def show_my_leaves(self):
        """Show user's own leaves."""
        dialog = tk.Toplevel(self.root)
        dialog.title("My Leaves")
        dialog.geometry("900x500")
        dialog.configure(bg=COLORS['background'])
        
        tk.Label(dialog, text="My Leave Requests", font=('Segoe UI', 16, 'bold'), bg=COLORS['background'], fg=COLORS['primary']).pack(pady=10)
        
        mine = [lv for lv in self.leaves if lv["username"] == self.current_user["username"]]
        if not mine:
            tk.Label(dialog, text="You have no leave requests.", font=('Segoe UI', 12), bg=COLORS['background'], fg=COLORS['text_light']).pack(pady=50)
            return
            
        tree_frame = tk.Frame(dialog, bg=COLORS['background'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=("ID", "Start Date", "End Date", "Type", "Status", "Reason"), show="headings", height=15)
        for col in ("ID", "Start Date", "End Date", "Type", "Status", "Reason"):
            tree.heading(col, text=col)
            tree.column(col, width=120)
        tree.column("Reason", width=250)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for lv in mine:
            tree.insert("", tk.END, values=(
                lv["leave_id"],
                lv["start_date"],
                lv["end_date"],
                lv["leave_type"],
                lv["status"],
                lv["reason"][:50] + "..." if len(lv["reason"]) > 50 else lv["reason"]
            ))
            
    def show_yearly_leaves(self):
        """Show yearly holidays/leaves."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Yearly Leaves / Holidays")
        dialog.geometry("800x600")
        dialog.configure(bg=COLORS['background'])
        
        # Title
        title_frame = tk.Frame(dialog, bg=COLORS['primary'], height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(
            title_frame,
            text="Yearly Leaves / Holidays",
            font=('Segoe UI', 18, 'bold'),
            bg=COLORS['primary'],
            fg=COLORS['surface'],
            pady=15
        ).pack()
        
        # Table frame
        table_frame = tk.Frame(dialog, bg=COLORS['background'])
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Create treeview with styled columns
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                    background=COLORS['surface'],
                    foreground=COLORS['text'],
                    rowheight=25,
                    fieldbackground=COLORS['surface'],
                    font=('Segoe UI', 10))
        style.configure("Treeview.Heading",
                    background=COLORS['success'],
                    foreground=COLORS['surface'],
                    font=('Segoe UI', 11, 'bold'),
                    relief=tk.FLAT)
        style.map("Treeview.Heading",
                background=[('active', COLORS['success'])])
        
        tree = ttk.Treeview(
            table_frame,
            columns=("Sr. No", "Date", "Day", "Occasion"),
            show="headings",
            height=20
        )
        
        # Configure columns
        tree.heading("Sr. No", text="Sr. No")
        tree.heading("Date", text="Date")
        tree.heading("Day", text="Day")
        tree.heading("Occasion", text="Occasion")
        
        tree.column("Sr. No", width=80, anchor=tk.CENTER)
        tree.column("Date", width=150, anchor=tk.CENTER)
        tree.column("Day", width=120, anchor=tk.CENTER)
        tree.column("Occasion", width=350, anchor=tk.W)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Format dates and populate tree
        if not self.holidays:
            tk.Label(
                dialog,
                text="No holidays configured.",
                font=('Segoe UI', 12),
                bg=COLORS['background'],
                fg=COLORS['text_light']
            ).pack(pady=50)
        else:
            for holiday in self.holidays:
                # Format date from YYYY-MM-DD to DD-Mon-YY or DD Month YYYY
                date_str = holiday.get("date", "")
                formatted_date = date_str
                try:
                    if date_str:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                        # Format as DD-Mon-YY (e.g., 26-Jan-26)
                        formatted_date = date_obj.strftime("%d-%b-%y")
                except:
                    pass
                
                tree.insert("", tk.END, values=(
                    holiday.get("sr_no", ""),
                    formatted_date,
                    holiday.get("day", ""),
                    holiday.get("occasion", "")
                ))
        
        # Close button
        btn_frame = tk.Frame(dialog, bg=COLORS['background'])
        btn_frame.pack(pady=15)
        
        tk.Button(
            btn_frame,
            text="Close",
            font=('Segoe UI', 11),
            bg=COLORS['secondary'],
            fg=COLORS['surface'],
            activebackground=COLORS['accent'],
            activeforeground=COLORS['surface'],
            relief=tk.FLAT,
            cursor='hand2',
            command=dialog.destroy,
            padx=25,
            pady=8
        ).pack()
        
    def show_cancel_leave_dialog(self):
        """Show dialog to cancel leave."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Cancel Leave")
        dialog.geometry("700x400")
        dialog.configure(bg=COLORS['background'])
        
        tk.Label(dialog, text="Cancel Leave Request", font=('Segoe UI', 16, 'bold'), bg=COLORS['background'], fg=COLORS['primary']).pack(pady=10)
        
        mine = [lv for lv in self.leaves if lv["username"] == self.current_user["username"]]
        cancellable = [lv for lv in mine if lv["status"] in ("PENDING", "APPROVED")]
        if not cancellable:
            tk.Label(dialog, text="You have no cancellable leave requests.", font=('Segoe UI', 12), bg=COLORS['background'], fg=COLORS['text_light']).pack(pady=50)
            return
            
        tree_frame = tk.Frame(dialog, bg=COLORS['background'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=("ID", "Start Date", "End Date", "Status", "Reason"), show="headings", height=10)
        for col in ("ID", "Start Date", "End Date", "Status", "Reason"):
            tree.heading(col, text=col)
            tree.column(col, width=120)
        tree.column("Reason", width=250)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for lv in cancellable:
            tree.insert("", tk.END, values=(
                lv["leave_id"],
                lv["start_date"],
                lv["end_date"],
                lv["status"],
                lv["reason"][:50] + "..." if len(lv["reason"]) > 50 else lv["reason"]
            ))
            
        def cancel_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a leave request.")
                return
            item = tree.item(selected[0])
            leave_id = int(item['values'][0])
            target = next(
                (lv for lv in self.leaves if lv["leave_id"] == leave_id and lv["username"] == self.current_user["username"]),
                None
            )
            if target and target["status"] in ("PENDING", "APPROVED"):
                target["status"] = "CANCELED"
                target["updated_at"] = datetime.now().isoformat(sep=" ", timespec="seconds")
                if self.save_data():
                    messagebox.showinfo("Success", f"Leave {leave_id} canceled.")
                    dialog.destroy()
                    self.show_user_dashboard()
            else:
                messagebox.showerror("Error", "Leave not found or cannot be canceled.")
                
        tk.Button(
            dialog,
            text="Cancel Selected Leave",
            font=('Segoe UI', 10),
            bg=COLORS['warning'],
            fg=COLORS['surface'],
            command=cancel_selected,
            padx=15,
            pady=5
        ).pack(pady=10)
        
    def show_master_creation_dialog(self):
        """Show dialog to create master user if none exists."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Create Master User")
        dialog.geometry("450x400")
        dialog.configure(bg=COLORS['surface'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"450x400+{x}+{y}")
        
        tk.Label(dialog, text="Create Master User", font=('Segoe UI', 16, 'bold'), bg=COLORS['surface'], fg=COLORS['primary']).pack(pady=20)
        tk.Label(dialog, text="No master user found. Please create one.", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=10)
        
        tk.Label(dialog, text="Master Username:", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        username_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30)
        username_entry.pack(pady=5)
        
        tk.Label(dialog, text="Full Name (optional):", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        fullname_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30)
        fullname_entry.pack(pady=5)
        
        tk.Label(dialog, text="Password:", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        password_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30, show="*")
        password_entry.pack(pady=5)
        
        tk.Label(dialog, text="Confirm Password:", font=('Segoe UI', 10), bg=COLORS['surface'], fg=COLORS['text']).pack(pady=5)
        confirm_entry = tk.Entry(dialog, font=('Segoe UI', 10), width=30, show="*")
        confirm_entry.pack(pady=5)
        
        def create_master():
            username = username_entry.get().strip()
            full_name = fullname_entry.get().strip()
            password = password_entry.get()
            confirm = confirm_entry.get()
            
            if not username:
                messagebox.showerror("Error", "Username cannot be empty.")
                return
            if username in self.users:
                messagebox.showerror("Error", "Username already exists.")
                return
            if password != confirm:
                messagebox.showerror("Error", "Passwords do not match.")
                return
            if not password:
                messagebox.showerror("Error", "Password cannot be empty.")
                return
                
            self.users[username] = {
                "username": username,
                "full_name": full_name,
                "password_hash": hash_password(password),
                "role": "master",
                "is_active": True,
            }
            if self.save_data():
                messagebox.showinfo("Success", "Master user created successfully.")
                dialog.destroy()
                
        tk.Button(
            dialog,
            text="Create Master",
            font=('Segoe UI', 11, 'bold'),
            bg=COLORS['success'],
            fg=COLORS['surface'],
            command=create_master,
            padx=20,
            pady=5
        ).pack(pady=20)
        
    def logout(self):
        """Handle logout."""
        self.current_user = None
        self.show_login_screen()
        
    def clear_main_frame(self):
        """Clear all widgets from main frame."""
        for widget in self.main_frame.winfo_children():
            widget.destroy()


def main():
    root = tk.Tk()
    app = LeaveTrackerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()